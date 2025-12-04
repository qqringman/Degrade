#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
JIRA Degrade % 分析系統 - 修復版
修復內容：
1. 解決合併數量與分開數量不一致的問題
2. 修正週次日期範圍計算，確保與 JIRA 查詢一致
3. 匯出 HTML 加入圖表顯示筆數和 Assignee 詳細分布表格
4. Degrade issues 使用 created 日期
5. Resolved issues 使用 created 日期
6. 趨勢圖加入 CCC issue 數量線
"""

from flask import Flask, jsonify, render_template, request, send_file
from flask_cors import CORS
from dotenv import load_dotenv
import os
from datetime import datetime, timedelta
from collections import defaultdict
from jira_degrade_manager import JiraDegradeManagerFast, load_all_filters_parallel
import time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import json
import socket

# 載入環境變數
load_dotenv()

app = Flask(__name__)
CORS(app)

# JIRA 連線設定
JIRA_CONFIG = {
    'internal': {
        'site': os.getenv('JIRA_SITE', 'jira.realtek.com'),
        'user': os.getenv('JIRA_USER'),
        'password': os.getenv('JIRA_PASSWORD'),
        'token': os.getenv('JIRA_TOKEN')
    },
    'vendor': {
        'site': os.getenv('VENDOR_JIRA_SITE', 'vendorjira.realtek.com'),
        'user': os.getenv('VENDOR_JIRA_USER'),
        'password': os.getenv('VENDOR_JIRA_PASSWORD'),
        'token': os.getenv('VENDOR_JIRA_TOKEN')
    }
}

# Filter IDs - 優先從 .env 讀取，沒有才使用預設值
FILTERS = {
    'degrade': {
        'internal': os.getenv('FILTER_INTERNAL_DEGRADE'),  # 內部 SQA+QC degrade from 2020/09/02
        'vendor': os.getenv('FILTER_VENDOR_DEGRADE')      # Vendor Jira QC Degrade from 2022/09/02
    },
    'resolved': {
        'internal': os.getenv('FILTER_INTERNAL_RESOLVED'),  # 內部 all resolved from 2020/09/02
        'vendor': os.getenv('FILTER_VENDOR_RESOLVED')       # Vendor all customer resolved from 2020/09/02
    }
}

# MTTR Filter IDs - 若有設定才啟用 MTTR 指標頁籤
MTTR_FILTERS = {
    'resolved': {  # 已解掉的問題
        'internal': os.getenv('FILTER_MTTR_INTERNAL_RESOLVED'),  # 內部 jira resolved issue
        'vendor': os.getenv('FILTER_MTTR_VENDOR_RESOLVED')       # Vendor resolved jira
    },
    'open': {  # 尚未解掉的問題
        'internal': os.getenv('FILTER_MTTR_INTERNAL_OPEN'),  # 內部 jira open issue
        'vendor': os.getenv('FILTER_MTTR_VENDOR_OPEN')       # Vendor open jira
    }
}

# 檢查是否啟用 MTTR 功能
MTTR_ENABLED = any([
    MTTR_FILTERS['resolved']['internal'],
    MTTR_FILTERS['resolved']['vendor'],
    MTTR_FILTERS['open']['internal'],
    MTTR_FILTERS['open']['vendor']
])

class DataCache:
    """記憶體快取"""
    def __init__(self, ttl_seconds=3600):
        self.data = None
        self.timestamp = None
        self.ttl = ttl_seconds
    
    def set(self, data):
        self.data = data
        self.timestamp = time.time()
    
    def get(self):
        if self.data is None or self.timestamp is None:
            return None
        
        if time.time() - self.timestamp > self.ttl:
            return None
        
        return self.data
    
    def age(self):
        """回傳快取年齡（秒）"""
        if self.timestamp is None:
            return None
        return time.time() - self.timestamp
    
    def clear(self):
        self.data = None
        self.timestamp = None

# 建立全域快取 - 從 .env 讀取 TTL（預設 1 小時）
cache = DataCache(ttl_seconds=int(os.getenv('CACHE_TTL', 3600)))

def load_data():
    """載入資料並快取"""
    try:
        print("�� 開始載入資料...")
        raw_data = load_all_filters_parallel(JIRA_CONFIG, FILTERS)
        
        # 驗證資料格式
        if not isinstance(raw_data, dict):
            print(f"❌ 錯誤: raw_data 不是字典，類型為 {type(raw_data)}")
            return None
        
        if 'degrade' not in raw_data or 'resolved' not in raw_data:
            print(f"❌ 錯誤: raw_data 缺少必要的鍵")
            print(f"   raw_data 的鍵: {list(raw_data.keys())}")
            return None
        
        # 檢查是否為新格式（包含 issues 子鍵）
        if isinstance(raw_data['degrade'], dict) and 'issues' in raw_data['degrade']:
            print("�� 檢測到新格式資料（包含統計資訊）")
            # 新格式：{'degrade': {'issues': [...], 'total': ..., 'weekly': ..., 'assignees': ...}}
            data = {
                'degrade': raw_data['degrade']['issues'],
                'resolved': raw_data['resolved']['issues'],
                'metadata': raw_data.get('metadata', {})
            }
            print(f"✅ 資料載入成功: degrade={len(data['degrade'])}, resolved={len(data['resolved'])}")
        elif isinstance(raw_data['degrade'], list):
            print("�� 檢測到舊格式資料（純列表）")
            # 舊格式：{'degrade': [...], 'resolved': [...]}
            data = raw_data
            print(f"✅ 資料載入成功: degrade={len(data['degrade'])}, resolved={len(data['resolved'])}")
        else:
            print(f"❌ 錯誤: data['degrade'] 格式不正確，類型為 {type(raw_data['degrade'])}")
            if isinstance(raw_data['degrade'], dict):
                print(f"   degrade 的鍵: {list(raw_data['degrade'].keys())}")
            return None
        
        # 驗證最終格式
        if not isinstance(data['degrade'], list):
            print(f"❌ 錯誤: 處理後 data['degrade'] 仍不是列表")
            return None
        
        if not isinstance(data['resolved'], list):
            print(f"❌ 錯誤: 處理後 data['resolved'] 仍不是列表")
            return None
        
        data['jira_sites'] = {
            'internal': JIRA_CONFIG['internal']['site'],
            'vendor': JIRA_CONFIG['vendor']['site']
        }
        cache.set(data)
        return data
    except Exception as e:
        print(f"❌ 載入資料失敗: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_data():
    """取得資料（優先使用快取）"""
    data = cache.get()
    if data:
        age = cache.age()
        print(f"✓ 使用快取 (年齡: {age:.0f}秒)")
        return data
    
    print("⚠ 快取無效，重新載入...")
    return load_data()

def get_iso_week_dates(year, week):
    """
    根據 ISO 8601 標準計算指定年份和週次的起始和結束日期
    修正：結束日期使用 23:59:59，確保包含當天所有時間
    """
    # 找到該年的第一天
    jan_4 = datetime(year, 1, 4)  # ISO 規則：包含 1 月 4 日的週就是第一週
    # 找到該週的星期一
    week_1_monday = jan_4 - timedelta(days=jan_4.weekday())
    # 計算目標週的星期一
    target_monday = week_1_monday + timedelta(weeks=week - 1)
    # 計算星期日（設定為 23:59:59）
    target_sunday = target_monday + timedelta(days=6, hours=23, minutes=59, seconds=59)
    
    return target_monday, target_sunday

def analyze_by_week_with_dates(issues, date_field='created'):
    """
    統計週次分布，並返回每週的起始和結束日期（符合 ISO 8601 標準）
    修正：準確計算週次邊界，包含整天的 issues
    
    Args:
        issues: issue 列表
        date_field: 要使用的日期欄位（'created' 或 'created'）
    """
    weekly_stats = {}
    
    for issue in issues:
        fields = issue.get('fields', {})
        date_str = fields.get(date_field)
        
        if not date_str:
            continue
        
        try:
            # 解析日期（可能包含時間）
            if 'T' in date_str:
                # 完整的 ISO 格式：2025-08-10T14:30:00.000+0800
                issue_date = datetime.fromisoformat(date_str.replace('Z', '+00:00').split('.')[0])
            else:
                # 只有日期：2025-08-10
                issue_date = datetime.strptime(date_str[:10], '%Y-%m-%d')
            
            # 計算 ISO 週次
            iso_calendar = issue_date.isocalendar()
            iso_year = iso_calendar[0]
            iso_week = iso_calendar[1]
            week_key = f"{iso_year}-W{iso_week:02d}"
            
            if week_key not in weekly_stats:
                # 使用正確的 ISO 週次計算方法
                week_start, week_end = get_iso_week_dates(iso_year, iso_week)
                
                weekly_stats[week_key] = {
                    'count': 0,
                    'issues': [],
                    'start_date': week_start.strftime('%Y-%m-%d'),
                    'end_date': week_end.strftime('%Y-%m-%d'),
                    # 新增：用於 JIRA JQL 查詢的精確時間
                    'start_datetime': week_start.strftime('%Y-%m-%d %H:%M'),
                    'end_datetime': week_end.strftime('%Y-%m-%d %H:%M')
                }
            
            weekly_stats[week_key]['count'] += 1
            weekly_stats[week_key]['issues'].append(issue.get('key'))
            
        except Exception as e:
            print(f"⚠️  週次統計錯誤: {e} (issue: {issue.get('key')}, date: {date_str})")
            continue
    
    return weekly_stats

def calculate_weekly_percentage(degrade_weekly, resolved_weekly):
    """計算每週百分比，並加入 degrade 和 resolved 的實際數量"""
    all_weeks = sorted(set(list(degrade_weekly.keys()) + list(resolved_weekly.keys())))
    
    weekly_stats = []
    for week in all_weeks:
        degrade_count = degrade_weekly.get(week, {}).get('count', 0)
        resolved_count = resolved_weekly.get(week, {}).get('count', 0)
        percentage = (degrade_count / resolved_count * 100) if resolved_count > 0 else 0
        
        weekly_stats.append({
            'week': week,
            'degrade_count': degrade_count,
            'resolved_count': resolved_count,
            'percentage': round(percentage, 2)
        })
    
    return weekly_stats

def filter_issues(issues, start_date, end_date, owner, date_field='created'):
    """
    過濾 issues
    
    Args:
        issues: issue 列表
        start_date: 開始日期
        end_date: 結束日期
        owner: Assignee 名稱
        date_field: 要使用的日期欄位（'created' 或 'created'）
    """
    filtered = []
    
    # 確保 issues 是列表
    if not isinstance(issues, list):
        print(f"⚠️  警告: issues 不是列表，類型為 {type(issues)}")
        return []
    
    for issue in issues:
        # 確保 issue 是字典
        if not isinstance(issue, dict):
            print(f"⚠️  警告: issue 不是字典，類型為 {type(issue)}")
            continue
            
        fields = issue.get('fields', {})
        
        # 日期過濾 - 使用指定的日期欄位
        if start_date or end_date:
            date_value = fields.get(date_field)
            if date_value:
                try:
                    # 解析日期（處理時間部分）
                    if 'T' in date_value:
                        issue_date = datetime.fromisoformat(date_value.replace('Z', '+00:00').split('.')[0])
                    else:
                        issue_date = datetime.strptime(date_value[:10], '%Y-%m-%d')
                    
                    if start_date:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        if issue_date < start_dt:
                            continue
                    
                    if end_date:
                        # 結束日期包含整天：23:59:59
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(hours=23, minutes=59, seconds=59)
                        if issue_date > end_dt:
                            continue
                except Exception as e:
                    print(f"⚠️  日期解析錯誤: {e} (issue: {issue.get('key')}, date: {date_value})")
                    pass
        
        # Owner 過濾
        if owner:
            assignee = fields.get('assignee')
            if isinstance(assignee, dict):
                assignee_name = assignee.get('displayName', 'Unassigned')
            else:
                assignee_name = 'Unassigned'
            
            if assignee_name != owner:
                continue
        
        filtered.append(issue)
    
    return filtered

@app.route('/')
def index():
    """首頁"""
    return render_template('index.html')

@app.route('/api/stats')
def get_stats():
    """取得統計資料"""
    try:
        data = get_data()
        if not data:
            return jsonify({'success': False, 'error': '載入資料失敗'}), 500
        
        # 確保資料格式正確
        if not isinstance(data.get('degrade'), list):
            print(f"❌ 錯誤: data['degrade'] 不是列表，類型為 {type(data.get('degrade'))}")
            return jsonify({'success': False, 'error': 'degrade 資料格式錯誤'}), 500
        
        if not isinstance(data.get('resolved'), list):
            print(f"❌ 錯誤: data['resolved'] 不是列表，類型為 {type(data.get('resolved'))}")
            return jsonify({'success': False, 'error': 'resolved 資料格式錯誤'}), 500
        
        # 取得過濾參數
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner')
        
        print(f"�� 過濾參數: start_date={start_date}, end_date={end_date}, owner={owner}")
        print(f"�� 原始資料: degrade={len(data['degrade'])}, resolved={len(data['resolved'])}")
        
        # 過濾資料 - degrade 使用 created，resolved 使用 created
        filtered_degrade = filter_issues(data['degrade'], start_date, end_date, owner, date_field='created')
        filtered_resolved = filter_issues(data['resolved'], start_date, end_date, owner, date_field='created')
        
        print(f"�� 過濾後: degrade={len(filtered_degrade)}, resolved={len(filtered_resolved)}")
        
        # 確保所有 issues 都有正確的 _source 標記
        missing_degrade = [i for i in filtered_degrade if i.get('_source') not in ['internal', 'vendor']]
        missing_resolved = [i for i in filtered_resolved if i.get('_source') not in ['internal', 'vendor']]
        
        if missing_degrade:
            print(f"⚠️  警告: 有 {len(missing_degrade)} 個 degrade issues 沒有正確的 _source 標記，正在修復...")
            for issue in missing_degrade:
                if 'vendorjira' in issue.get('self', '').lower():
                    issue['_source'] = 'vendor'
                else:
                    issue['_source'] = 'internal'
        
        if missing_resolved:
            print(f"⚠️  警告: 有 {len(missing_resolved)} 個 resolved issues 沒有正確的 _source 標記，正在修復...")
            for issue in missing_resolved:
                if 'vendorjira' in issue.get('self', '').lower():
                    issue['_source'] = 'vendor'
                else:
                    issue['_source'] = 'internal'
        
        # 分離內部和 Vendor
        internal_degrade = [i for i in filtered_degrade if i.get('_source') == 'internal']
        vendor_degrade = [i for i in filtered_degrade if i.get('_source') == 'vendor']
        internal_resolved = [i for i in filtered_resolved if i.get('_source') == 'internal']
        vendor_resolved = [i for i in filtered_resolved if i.get('_source') == 'vendor']
        
        # 驗證數量一致性
        print(f"�� 分離驗證:")
        print(f"   Degrade: total={len(filtered_degrade)}, internal={len(internal_degrade)}, vendor={len(vendor_degrade)}, sum={len(internal_degrade)+len(vendor_degrade)}")
        print(f"   Resolved: total={len(filtered_resolved)}, internal={len(internal_resolved)}, vendor={len(vendor_resolved)}, sum={len(internal_resolved)+len(vendor_resolved)}")
        
        # 收集所有 assignees
        all_owners = set()
        for issue in data['degrade'] + data['resolved']:
            fields = issue.get('fields', {})
            assignee = fields.get('assignee')
            if assignee:
                all_owners.add(assignee.get('displayName', 'Unassigned'))
            else:
                all_owners.add('Unassigned')
        
        # 重新統計
        manager = JiraDegradeManagerFast(
            site=JIRA_CONFIG['internal']['site'],
            user=JIRA_CONFIG['internal']['user'],
            password=JIRA_CONFIG['internal']['password'],
            token=JIRA_CONFIG['internal']['token']
        )
        
        total_degrade = len(filtered_degrade)
        total_resolved = len(filtered_resolved)
        overall_percentage = (total_degrade / total_resolved * 100) if total_resolved > 0 else 0
        
        # Assignee 分布
        degrade_assignees_internal = manager.get_assignee_distribution(internal_degrade)
        degrade_assignees_vendor = manager.get_assignee_distribution(vendor_degrade)
        resolved_assignees_internal = manager.get_assignee_distribution(internal_resolved)
        resolved_assignees_vendor = manager.get_assignee_distribution(vendor_resolved)
        
        # 週次統計 - degrade 使用 created，resolved 使用 created
        degrade_weekly = analyze_by_week_with_dates(filtered_degrade, date_field='created')
        resolved_weekly = analyze_by_week_with_dates(filtered_resolved, date_field='created')
        weekly_stats = calculate_weekly_percentage(degrade_weekly, resolved_weekly)
        
        degrade_weekly_internal = analyze_by_week_with_dates(internal_degrade, date_field='created')
        degrade_weekly_vendor = analyze_by_week_with_dates(vendor_degrade, date_field='created')
        resolved_weekly_internal = analyze_by_week_with_dates(internal_resolved, date_field='created')
        resolved_weekly_vendor = analyze_by_week_with_dates(vendor_resolved, date_field='created')
        
        return jsonify({
            'success': True,
            'data': {
                'overall': {
                    'degrade_count': total_degrade,
                    'resolved_count': total_resolved,
                    'percentage': round(overall_percentage, 2),
                    'internal': {
                        'degrade_count': len(internal_degrade),
                        'resolved_count': len(internal_resolved)
                    },
                    'vendor': {
                        'degrade_count': len(vendor_degrade),
                        'resolved_count': len(vendor_resolved)
                    }
                },
                'weekly': weekly_stats,
                'weekly_by_source': {
                    'internal': degrade_weekly_internal,
                    'vendor': degrade_weekly_vendor
                },
                'weekly_by_source_resolved': {
                    'internal': resolved_weekly_internal,
                    'vendor': resolved_weekly_vendor
                },
                'assignees': {
                    'degrade': {
                        'internal': degrade_assignees_internal,
                        'vendor': degrade_assignees_vendor
                    },
                    'resolved': {
                        'internal': resolved_assignees_internal,
                        'vendor': resolved_assignees_vendor
                    }
                },
                'jira_sites': data['jira_sites'],
                'all_owners': sorted(list(all_owners)),
                'filters': {
                    'start_date': start_date,
                    'end_date': end_date,
                    'owner': owner
                },
                'filter_ids': FILTERS,
                'cache_age': cache.age(),
                'load_time': data['metadata']['load_time'],
                'warnings': data['metadata'].get('warnings', [])
            }
        })
        
    except Exception as e:
        print(f"❌ API 錯誤: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# ===== MTTR 相關功能 =====

# MTTR 專用快取
mttr_cache = DataCache(ttl_seconds=int(os.getenv('CACHE_TTL', 3600)))

def load_mttr_data():
    """載入 MTTR 資料"""
    if not MTTR_ENABLED:
        return None

    try:
        print("�� 開始載入 MTTR 資料...")
        from jira_degrade_manager import JiraDegradeManagerFast

        # 建立 JIRA managers
        internal_jira = JiraDegradeManagerFast(
            site=JIRA_CONFIG['internal']['site'],
            user=JIRA_CONFIG['internal']['user'],
            password=JIRA_CONFIG['internal']['password'],
            token=JIRA_CONFIG['internal']['token']
        )

        vendor_jira = JiraDegradeManagerFast(
            site=JIRA_CONFIG['vendor']['site'],
            user=JIRA_CONFIG['vendor']['user'],
            password=JIRA_CONFIG['vendor']['password'],
            token=JIRA_CONFIG['vendor']['token']
        )

        results = {}
        warnings = []
        start_time = time.time()

        # 定義要載入的任務
        tasks = []
        if MTTR_FILTERS['resolved']['internal']:
            tasks.append(('resolved_internal', internal_jira, MTTR_FILTERS['resolved']['internal'], 'internal', 'resolved'))
        if MTTR_FILTERS['resolved']['vendor']:
            tasks.append(('resolved_vendor', vendor_jira, MTTR_FILTERS['resolved']['vendor'], 'vendor', 'resolved'))
        if MTTR_FILTERS['open']['internal']:
            tasks.append(('open_internal', internal_jira, MTTR_FILTERS['open']['internal'], 'internal', 'open'))
        if MTTR_FILTERS['open']['vendor']:
            tasks.append(('open_vendor', vendor_jira, MTTR_FILTERS['open']['vendor'], 'vendor', 'open'))

        # 並行載入
        from concurrent.futures import ThreadPoolExecutor, as_completed

        with ThreadPoolExecutor(max_workers=4) as executor:
            future_to_task = {
                executor.submit(jira.get_filter_issues_fast, filter_id): (task_name, source, type_name)
                for task_name, jira, filter_id, source, type_name in tasks
            }

            for future in as_completed(future_to_task):
                task_name, source, type_name = future_to_task[future]
                try:
                    result = future.result()

                    if result['success']:
                        # 標記來源
                        for issue in result['issues']:
                            issue['_source'] = source
                        results[task_name] = result['issues']
                    else:
                        results[task_name] = []
                        warnings.append({
                            'source': source,
                            'type': type_name,
                            'site': result.get('site', ''),
                            'filter_id': result.get('filter_id', ''),
                            'filter_owner': result.get('filter_owner', 'Unknown'),
                            'error': result.get('error', '未知錯誤'),
                            'error_type': result.get('error_type', 'UNKNOWN_ERROR')
                        })
                except Exception as e:
                    print(f"  ❌ MTTR {task_name} 失敗: {e}")
                    results[task_name] = []

        total_time = time.time() - start_time
        print(f"✅ MTTR 資料載入完成！耗時: {total_time:.1f} 秒")

        data = {
            'resolved': {
                'internal': results.get('resolved_internal', []),
                'vendor': results.get('resolved_vendor', [])
            },
            'open': {
                'internal': results.get('open_internal', []),
                'vendor': results.get('open_vendor', [])
            },
            'jira_sites': {
                'internal': JIRA_CONFIG['internal']['site'],
                'vendor': JIRA_CONFIG['vendor']['site']
            },
            'metadata': {
                'load_time': total_time,
                'timestamp': datetime.now().isoformat(),
                'warnings': warnings
            }
        }

        mttr_cache.set(data)
        return data

    except Exception as e:
        print(f"❌ MTTR 載入資料失敗: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_mttr_data():
    """取得 MTTR 資料（優先使用快取）"""
    data = mttr_cache.get()
    if data:
        age = mttr_cache.age()
        print(f"✓ 使用 MTTR 快取 (年齡: {age:.0f}秒)")
        return data

    print("⚠ MTTR 快取無效，重新載入...")
    return load_mttr_data()

def calculate_mttr_metrics(issues, metric_type='resolved'):
    """
    計算 MTTR 指標

    Args:
        issues: issue 列表
        metric_type: 'resolved' (已解掉) 或 'open' (未解掉)

    Returns:
        dict: 包含 MTTR 統計資料
    """
    weekly_stats = {}
    now = datetime.now()

    for issue in issues:
        fields = issue.get('fields', {})
        created_str = fields.get('created')

        if not created_str:
            continue

        try:
            # 解析 created 日期
            if 'T' in created_str:
                created_date = datetime.fromisoformat(created_str.replace('Z', '+00:00').split('.')[0])
            else:
                created_date = datetime.strptime(created_str[:10], '%Y-%m-%d')

            # 計算 ISO 週次
            iso_calendar = created_date.isocalendar()
            iso_year = iso_calendar[0]
            iso_week = iso_calendar[1]
            week_key = f"{iso_year}-W{iso_week:02d}"

            if week_key not in weekly_stats:
                week_start, week_end = get_iso_week_dates(iso_year, iso_week)
                weekly_stats[week_key] = {
                    'count': 0,
                    'total_mttr_days': 0,
                    'mttr_issues': [],
                    'overdue_count': 0,
                    'total_overdue_days': 0,
                    'overdue_issues': [],
                    'start_date': week_start.strftime('%Y-%m-%d'),
                    'end_date': week_end.strftime('%Y-%m-%d')
                }

            weekly_stats[week_key]['count'] += 1

            if metric_type == 'resolved':
                # 已解掉的問題: MTTR = Resolved - Created
                resolved_str = fields.get('resolutiondate')
                if resolved_str:
                    if 'T' in resolved_str:
                        resolved_date = datetime.fromisoformat(resolved_str.replace('Z', '+00:00').split('.')[0])
                    else:
                        resolved_date = datetime.strptime(resolved_str[:10], '%Y-%m-%d')

                    mttr_days = (resolved_date - created_date).days
                    weekly_stats[week_key]['total_mttr_days'] += mttr_days
                    weekly_stats[week_key]['mttr_issues'].append({
                        'key': issue.get('key'),
                        'mttr_days': mttr_days
                    })

                    # Overdue: Resolved - Duedate (duedate 為空則不計)
                    duedate_str = fields.get('duedate')
                    if duedate_str:
                        if 'T' in duedate_str:
                            duedate = datetime.fromisoformat(duedate_str.replace('Z', '+00:00').split('.')[0])
                        else:
                            duedate = datetime.strptime(duedate_str[:10], '%Y-%m-%d')

                        overdue_days = (resolved_date - duedate).days
                        if overdue_days > 0:
                            weekly_stats[week_key]['overdue_count'] += 1
                            weekly_stats[week_key]['total_overdue_days'] += overdue_days
                            weekly_stats[week_key]['overdue_issues'].append({
                                'key': issue.get('key'),
                                'overdue_days': overdue_days
                            })
            else:
                # 尚未解掉的問題: MTTR(ongoing) = Now - Created
                mttr_days = (now - created_date).days
                weekly_stats[week_key]['total_mttr_days'] += mttr_days
                weekly_stats[week_key]['mttr_issues'].append({
                    'key': issue.get('key'),
                    'mttr_days': mttr_days
                })

                # Overdue(ongoing): Now - Duedate (duedate 為空則不計)
                duedate_str = fields.get('duedate')
                if duedate_str:
                    if 'T' in duedate_str:
                        duedate = datetime.fromisoformat(duedate_str.replace('Z', '+00:00').split('.')[0])
                    else:
                        duedate = datetime.strptime(duedate_str[:10], '%Y-%m-%d')

                    overdue_days = (now - duedate).days
                    if overdue_days > 0:
                        weekly_stats[week_key]['overdue_count'] += 1
                        weekly_stats[week_key]['total_overdue_days'] += overdue_days
                        weekly_stats[week_key]['overdue_issues'].append({
                            'key': issue.get('key'),
                            'overdue_days': overdue_days
                        })

        except Exception as e:
            print(f"⚠️  MTTR 計算錯誤: {e} (issue: {issue.get('key')})")
            continue

    # 計算平均值
    result = []
    for week in sorted(weekly_stats.keys()):
        stats = weekly_stats[week]
        avg_mttr = stats['total_mttr_days'] / stats['count'] if stats['count'] > 0 else 0
        avg_overdue = stats['total_overdue_days'] / stats['overdue_count'] if stats['overdue_count'] > 0 else 0

        result.append({
            'week': week,
            'count': stats['count'],
            'avg_mttr_days': round(avg_mttr, 2),
            'overdue_count': stats['overdue_count'],
            'avg_overdue_days': round(avg_overdue, 2),
            'start_date': stats['start_date'],
            'end_date': stats['end_date']
        })

    return result

@app.route('/api/mttr/enabled')
def mttr_enabled():
    """檢查 MTTR 功能是否啟用"""
    return jsonify({
        'enabled': MTTR_ENABLED,
        'filters': MTTR_FILTERS if MTTR_ENABLED else None
    })

@app.route('/api/mttr/stats')
def get_mttr_stats():
    """取得 MTTR 統計資料"""
    if not MTTR_ENABLED:
        return jsonify({'success': False, 'error': 'MTTR 功能未啟用'}), 400

    try:
        data = get_mttr_data()
        if not data:
            return jsonify({'success': False, 'error': '載入 MTTR 資料失敗'}), 500

        # 取得過濾參數
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner')

        print(f"�� MTTR 過濾參數: start_date={start_date}, end_date={end_date}, owner={owner}")

        # 過濾資料 (使用 created 欄位)
        resolved_internal = filter_issues(data['resolved']['internal'], start_date, end_date, owner, date_field='created')
        resolved_vendor = filter_issues(data['resolved']['vendor'], start_date, end_date, owner, date_field='created')
        open_internal = filter_issues(data['open']['internal'], start_date, end_date, owner, date_field='created')
        open_vendor = filter_issues(data['open']['vendor'], start_date, end_date, owner, date_field='created')

        # 合併資料
        all_resolved = resolved_internal + resolved_vendor
        all_open = open_internal + open_vendor

        # 計算 MTTR 指標
        resolved_stats_internal = calculate_mttr_metrics(resolved_internal, 'resolved')
        resolved_stats_vendor = calculate_mttr_metrics(resolved_vendor, 'resolved')
        resolved_stats_all = calculate_mttr_metrics(all_resolved, 'resolved')

        open_stats_internal = calculate_mttr_metrics(open_internal, 'open')
        open_stats_vendor = calculate_mttr_metrics(open_vendor, 'open')
        open_stats_all = calculate_mttr_metrics(all_open, 'open')

        # 收集所有 assignees
        all_owners = set()
        for issues in [resolved_internal, resolved_vendor, open_internal, open_vendor]:
            for issue in issues:
                fields = issue.get('fields', {})
                assignee = fields.get('assignee')
                if assignee:
                    all_owners.add(assignee.get('displayName', 'Unassigned'))
                else:
                    all_owners.add('Unassigned')

        return jsonify({
            'success': True,
            'data': {
                'resolved': {
                    'all': resolved_stats_all,
                    'internal': resolved_stats_internal,
                    'vendor': resolved_stats_vendor,
                    'counts': {
                        'total': len(all_resolved),
                        'internal': len(resolved_internal),
                        'vendor': len(resolved_vendor)
                    }
                },
                'open': {
                    'all': open_stats_all,
                    'internal': open_stats_internal,
                    'vendor': open_stats_vendor,
                    'counts': {
                        'total': len(all_open),
                        'internal': len(open_internal),
                        'vendor': len(open_vendor)
                    }
                },
                'jira_sites': data['jira_sites'],
                'filter_ids': MTTR_FILTERS,
                'all_owners': sorted(list(all_owners)),
                'filters': {
                    'start_date': start_date,
                    'end_date': end_date,
                    'owner': owner
                },
                'cache_age': mttr_cache.age(),
                'load_time': data['metadata']['load_time'],
                'warnings': data['metadata'].get('warnings', [])
            }
        })

    except Exception as e:
        print(f"❌ MTTR API 錯誤: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mttr/refresh', methods=['POST'])
def refresh_mttr():
    """強制重新載入 MTTR 資料"""
    if not MTTR_ENABLED:
        return jsonify({'success': False, 'error': 'MTTR 功能未啟用'}), 400

    try:
        mttr_cache.clear()
        data = load_mttr_data()
        if data:
            return jsonify({'success': True, 'message': 'MTTR 資料重新載入完成'})
        else:
            return jsonify({'success': False, 'error': '載入失敗'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cache-status')
def cache_status():
    """快取狀態"""
    age = cache.age()
    mttr_age = mttr_cache.age() if MTTR_ENABLED else None
    return jsonify({
        'valid': age is not None and age < cache.ttl,
        'age_seconds': age,
        'age_minutes': age / 60 if age else None,
        'mttr_valid': mttr_age is not None and mttr_age < mttr_cache.ttl if MTTR_ENABLED else None,
        'mttr_age_seconds': mttr_age,
        'mttr_age_minutes': mttr_age / 60 if mttr_age else None
    })

@app.route('/api/refresh', methods=['POST'])
def refresh():
    """強制重新載入資料"""
    try:
        cache.clear()
        data = load_data()
        if data:
            return jsonify({'success': True, 'message': '資料重新載入完成'})
        else:
            return jsonify({'success': False, 'error': '載入失敗'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/excel')
def export_excel():
    """匯出 Excel - 多頁籤，可 filter"""
    try:
        data = get_data()
        if not data:
            return jsonify({'success': False, 'error': '無資料可匯出'}), 500
        
        # 取得過濾參數
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner')
        
        # 過濾資料
        filtered_degrade = filter_issues(data['degrade'], start_date, end_date, owner, date_field='created')
        filtered_resolved = filter_issues(data['resolved'], start_date, end_date, owner, date_field='created')
        
        # 建立 Excel
        wb = Workbook()
        wb.remove(wb.active)  # 移除預設工作表
        
        # 樣式定義
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def create_sheet(wb, title, data, columns, source_filter=None):
            """建立工作表"""
            ws = wb.create_sheet(title=title)
            
            # 過濾資料
            if source_filter:
                data = [i for i in data if i.get('_source') == source_filter]
            
            # 寫入標題
            for col_idx, (header, _) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 寫入資料
            for row_idx, issue in enumerate(data, 2):
                fields = issue.get('fields', {})
                for col_idx, (_, field_func) in enumerate(columns, 1):
                    value = field_func(issue, fields)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
            
            # 自動調整欄寬
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # 啟用篩選
            ws.auto_filter.ref = ws.dimensions
            
            return ws
        
        # 定義欄位 - degrade 使用 created，resolved 使用 created
        degrade_columns = [
            ('Issue Key', lambda i, f: i.get('key', '')),
            ('Assignee', lambda i, f: f.get('assignee', {}).get('displayName', 'Unassigned') if f.get('assignee') else 'Unassigned'),
            ('Created', lambda i, f: f.get('created', '')[:10] if f.get('created') else ''),
            ('Week', lambda i, f: f"{datetime.strptime(f.get('created', '')[:10], '%Y-%m-%d').isocalendar()[0]}-W{datetime.strptime(f.get('created', '')[:10], '%Y-%m-%d').isocalendar()[1]:02d}" if f.get('created') else ''),
            ('Source', lambda i, f: i.get('_source', 'unknown').upper())
        ]
        
        resolved_columns = [
            ('Issue Key', lambda i, f: i.get('key', '')),
            ('Assignee', lambda i, f: f.get('assignee', {}).get('displayName', 'Unassigned') if f.get('assignee') else 'Unassigned'),
            ('Resolved Date', lambda i, f: f.get('created', '')[:10] if f.get('created') else ''),
            ('Week', lambda i, f: f"{datetime.strptime(f.get('created', '')[:10], '%Y-%m-%d').isocalendar()[0]}-W{datetime.strptime(f.get('created', '')[:10], '%Y-%m-%d').isocalendar()[1]:02d}" if f.get('created') else ''),
            ('Source', lambda i, f: i.get('_source', 'unknown').upper())
        ]
        
        # 建立工作表
        create_sheet(wb, 'Degrade All', filtered_degrade, degrade_columns)
        create_sheet(wb, 'Degrade Internal', filtered_degrade, degrade_columns, 'internal')
        create_sheet(wb, 'Degrade Vendor', filtered_degrade, degrade_columns, 'vendor')
        create_sheet(wb, 'Resolved All', filtered_resolved, resolved_columns)
        create_sheet(wb, 'Resolved Internal', filtered_resolved, resolved_columns, 'internal')
        create_sheet(wb, 'Resolved Vendor', filtered_resolved, resolved_columns, 'vendor')
        
        # ===== MTTR 資料 =====
        mttr_summary_data = []
        if MTTR_ENABLED:
            mttr_data = get_mttr_data()
            if mttr_data:
                # 過濾 MTTR 資料
                mttr_resolved_internal = filter_issues(mttr_data['resolved']['internal'], start_date, end_date, owner, date_field='created')
                mttr_resolved_vendor = filter_issues(mttr_data['resolved']['vendor'], start_date, end_date, owner, date_field='created')
                mttr_open_internal = filter_issues(mttr_data['open']['internal'], start_date, end_date, owner, date_field='created')
                mttr_open_vendor = filter_issues(mttr_data['open']['vendor'], start_date, end_date, owner, date_field='created')

                mttr_all_resolved = mttr_resolved_internal + mttr_resolved_vendor
                mttr_all_open = mttr_open_internal + mttr_open_vendor

                # MTTR 欄位定義
                def calc_mttr_resolved(issue, fields):
                    """計算已解決問題的 MTTR"""
                    created = fields.get('created', '')
                    resolved = fields.get('resolutiondate', '')
                    if created and resolved:
                        try:
                            created_dt = datetime.fromisoformat(created.split('.')[0].replace('Z', ''))
                            resolved_dt = datetime.fromisoformat(resolved.split('.')[0].replace('Z', ''))
                            return (resolved_dt - created_dt).days
                        except:
                            pass
                    return ''

                def calc_overdue_resolved(issue, fields):
                    """計算已解決問題的 Overdue 天數"""
                    resolved = fields.get('resolutiondate', '')
                    duedate = fields.get('duedate', '')
                    if resolved and duedate:
                        try:
                            resolved_dt = datetime.fromisoformat(resolved.split('.')[0].replace('Z', ''))
                            due_dt = datetime.strptime(duedate[:10], '%Y-%m-%d')
                            days = (resolved_dt - due_dt).days
                            return days if days > 0 else 0
                        except:
                            pass
                    return ''

                def calc_mttr_open(issue, fields):
                    """計算未解決問題的 MTTR(ongoing)"""
                    created = fields.get('created', '')
                    if created:
                        try:
                            created_dt = datetime.fromisoformat(created.split('.')[0].replace('Z', ''))
                            return (datetime.now() - created_dt).days
                        except:
                            pass
                    return ''

                def calc_overdue_open(issue, fields):
                    """計算未解決問題的 Overdue 天數"""
                    duedate = fields.get('duedate', '')
                    if duedate:
                        try:
                            due_dt = datetime.strptime(duedate[:10], '%Y-%m-%d')
                            days = (datetime.now() - due_dt).days
                            return days if days > 0 else 0
                        except:
                            pass
                    return ''

                mttr_resolved_columns = [
                    ('Issue Key', lambda i, f: i.get('key', '')),
                    ('Assignee', lambda i, f: f.get('assignee', {}).get('displayName', 'Unassigned') if f.get('assignee') else 'Unassigned'),
                    ('Created', lambda i, f: f.get('created', '')[:10] if f.get('created') else ''),
                    ('Resolved', lambda i, f: f.get('resolutiondate', '')[:10] if f.get('resolutiondate') else ''),
                    ('Due Date', lambda i, f: f.get('duedate', '')[:10] if f.get('duedate') else ''),
                    ('MTTR (Days)', calc_mttr_resolved),
                    ('Overdue (Days)', calc_overdue_resolved),
                    ('Source', lambda i, f: i.get('_source', 'unknown').upper())
                ]

                mttr_open_columns = [
                    ('Issue Key', lambda i, f: i.get('key', '')),
                    ('Assignee', lambda i, f: f.get('assignee', {}).get('displayName', 'Unassigned') if f.get('assignee') else 'Unassigned'),
                    ('Created', lambda i, f: f.get('created', '')[:10] if f.get('created') else ''),
                    ('Due Date', lambda i, f: f.get('duedate', '')[:10] if f.get('duedate') else ''),
                    ('MTTR Ongoing (Days)', calc_mttr_open),
                    ('Overdue (Days)', calc_overdue_open),
                    ('Source', lambda i, f: i.get('_source', 'unknown').upper())
                ]

                # 建立 MTTR 工作表
                create_sheet(wb, 'MTTR Resolved All', mttr_all_resolved, mttr_resolved_columns)
                create_sheet(wb, 'MTTR Resolved Internal', mttr_resolved_internal, mttr_resolved_columns)
                create_sheet(wb, 'MTTR Resolved Vendor', mttr_resolved_vendor, mttr_resolved_columns)
                create_sheet(wb, 'MTTR Open All', mttr_all_open, mttr_open_columns)
                create_sheet(wb, 'MTTR Open Internal', mttr_open_internal, mttr_open_columns)
                create_sheet(wb, 'MTTR Open Vendor', mttr_open_vendor, mttr_open_columns)

                # MTTR 摘要資料
                mttr_summary_data = [
                    ['', ''],
                    ['===== MTTR 指標 =====', ''],
                    ['MTTR Resolved Issues (Total)', len(mttr_all_resolved)],
                    ['MTTR Resolved Issues (Internal)', len(mttr_resolved_internal)],
                    ['MTTR Resolved Issues (Vendor)', len(mttr_resolved_vendor)],
                    ['MTTR Open Issues (Total)', len(mttr_all_open)],
                    ['MTTR Open Issues (Internal)', len(mttr_open_internal)],
                    ['MTTR Open Issues (Vendor)', len(mttr_open_vendor)],
                ]

        # 統計摘要
        ws_summary = wb.create_sheet(title='Summary', index=0)
        summary_data = [
            ['統計項目', '數量'],
            ['===== Degrade 分析 =====', ''],
            ['Degrade Issues (Total)', len(filtered_degrade)],
            ['Degrade Issues (Internal)', len([i for i in filtered_degrade if i.get('_source') == 'internal'])],
            ['Degrade Issues (Vendor)', len([i for i in filtered_degrade if i.get('_source') == 'vendor'])],
            ['Resolved Issues (Total)', len(filtered_resolved)],
            ['Resolved Issues (Internal)', len([i for i in filtered_resolved if i.get('_source') == 'internal'])],
            ['Resolved Issues (Vendor)', len([i for i in filtered_resolved if i.get('_source') == 'vendor'])],
            ['Degrade %', f"{(len(filtered_degrade) / len(filtered_resolved) * 100) if len(filtered_resolved) > 0 else 0:.2f}%"],
        ] + mttr_summary_data + [
            ['', ''],
            ['===== 說明 =====', ''],
            ['Degrade Issues', '使用 created 日期'],
            ['Resolved Issues', '使用 created 日期'],
            ['MTTR', 'Resolved Date - Created Date'],
            ['MTTR (ongoing)', 'Now - Created Date'],
            ['Overdue', 'Resolved Date / Now - Due Date'],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
            cell_value = ws_summary.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                cell_label.font = header_font
                cell_label.fill = header_fill
                cell_value.font = header_font
                cell_value.fill = header_fill
            cell_label.border = border
            cell_value.border = border
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # 儲存到記憶體
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"jira_degrade_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Excel 匯出失敗: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/html')
def export_html():
    """匯出 HTML - 完整功能版，包含可點擊圖表和詳細表格，以及 MTTR 指標"""
    try:
        data = get_data()
        if not data:
            return jsonify({'success': False, 'error': '無資料可匯出'}), 500
        
        # 取得過濾參數
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner')
        chart_limit = int(request.args.get('chart_limit', 20))  # 圖表顯示筆數
        
        print(f"�� 匯出 HTML: chart_limit={chart_limit}, MTTR_ENABLED={MTTR_ENABLED}")

        # ===== MTTR 資料處理 =====
        mttr_html_section = ""
        mttr_js_section = ""

        if MTTR_ENABLED:
            mttr_data = get_mttr_data()
            if mttr_data:
                # 過濾 MTTR 資料
                mttr_resolved_internal = filter_issues(mttr_data['resolved']['internal'], start_date, end_date, owner, date_field='created')
                mttr_resolved_vendor = filter_issues(mttr_data['resolved']['vendor'], start_date, end_date, owner, date_field='created')
                mttr_open_internal = filter_issues(mttr_data['open']['internal'], start_date, end_date, owner, date_field='created')
                mttr_open_vendor = filter_issues(mttr_data['open']['vendor'], start_date, end_date, owner, date_field='created')

                mttr_all_resolved = mttr_resolved_internal + mttr_resolved_vendor
                mttr_all_open = mttr_open_internal + mttr_open_vendor

                # 計算 MTTR 指標
                mttr_resolved_stats_all = calculate_mttr_metrics(mttr_all_resolved, 'resolved')
                mttr_resolved_stats_internal = calculate_mttr_metrics(mttr_resolved_internal, 'resolved')
                mttr_resolved_stats_vendor = calculate_mttr_metrics(mttr_resolved_vendor, 'resolved')

                mttr_open_stats_all = calculate_mttr_metrics(mttr_all_open, 'open')
                mttr_open_stats_internal = calculate_mttr_metrics(mttr_open_internal, 'open')
                mttr_open_stats_vendor = calculate_mttr_metrics(mttr_open_vendor, 'open')

                # 準備圖表數據
                mttr_resolved_all_labels = json.dumps([w['week'] for w in mttr_resolved_stats_all])
                mttr_resolved_all_mttr = json.dumps([w['avg_mttr_days'] for w in mttr_resolved_stats_all])
                mttr_resolved_all_overdue = json.dumps([w['avg_overdue_days'] for w in mttr_resolved_stats_all])
                mttr_resolved_all_count = json.dumps([w['count'] for w in mttr_resolved_stats_all])

                mttr_resolved_internal_labels = json.dumps([w['week'] for w in mttr_resolved_stats_internal])
                mttr_resolved_internal_mttr = json.dumps([w['avg_mttr_days'] for w in mttr_resolved_stats_internal])
                mttr_resolved_internal_overdue = json.dumps([w['avg_overdue_days'] for w in mttr_resolved_stats_internal])
                mttr_resolved_internal_count = json.dumps([w['count'] for w in mttr_resolved_stats_internal])

                mttr_resolved_vendor_labels = json.dumps([w['week'] for w in mttr_resolved_stats_vendor])
                mttr_resolved_vendor_mttr = json.dumps([w['avg_mttr_days'] for w in mttr_resolved_stats_vendor])
                mttr_resolved_vendor_overdue = json.dumps([w['avg_overdue_days'] for w in mttr_resolved_stats_vendor])
                mttr_resolved_vendor_count = json.dumps([w['count'] for w in mttr_resolved_stats_vendor])

                mttr_open_all_labels = json.dumps([w['week'] for w in mttr_open_stats_all])
                mttr_open_all_mttr = json.dumps([w['avg_mttr_days'] for w in mttr_open_stats_all])
                mttr_open_all_overdue = json.dumps([w['avg_overdue_days'] for w in mttr_open_stats_all])
                mttr_open_all_count = json.dumps([w['count'] for w in mttr_open_stats_all])

                mttr_open_internal_labels = json.dumps([w['week'] for w in mttr_open_stats_internal])
                mttr_open_internal_mttr = json.dumps([w['avg_mttr_days'] for w in mttr_open_stats_internal])
                mttr_open_internal_overdue = json.dumps([w['avg_overdue_days'] for w in mttr_open_stats_internal])
                mttr_open_internal_count = json.dumps([w['count'] for w in mttr_open_stats_internal])

                mttr_open_vendor_labels = json.dumps([w['week'] for w in mttr_open_stats_vendor])
                mttr_open_vendor_mttr = json.dumps([w['avg_mttr_days'] for w in mttr_open_stats_vendor])
                mttr_open_vendor_overdue = json.dumps([w['avg_overdue_days'] for w in mttr_open_stats_vendor])
                mttr_open_vendor_count = json.dumps([w['count'] for w in mttr_open_stats_vendor])

                # 週次日期範圍
                mttr_resolved_internal_dates = json.dumps({w['week']: {'start_date': w['start_date'], 'end_date': w['end_date']} for w in mttr_resolved_stats_internal})
                mttr_resolved_vendor_dates = json.dumps({w['week']: {'start_date': w['start_date'], 'end_date': w['end_date']} for w in mttr_resolved_stats_vendor})
                mttr_open_internal_dates = json.dumps({w['week']: {'start_date': w['start_date'], 'end_date': w['end_date']} for w in mttr_open_stats_internal})
                mttr_open_vendor_dates = json.dumps({w['week']: {'start_date': w['start_date'], 'end_date': w['end_date']} for w in mttr_open_stats_vendor})

                # MTTR HTML 區塊
                mttr_html_section = f"""
        <!-- MTTR 指標區塊 -->
        <div class="info-banner">
            <strong>�� 提示：</strong> 圖表可以點擊！點擊內部 JIRA 或 Vendor JIRA 的圖表可跳轉到 JIRA 查看該週的 issues
        </div>

        <div class="stats-grid">
            <div class="stat-card">
                <h3>Resolved Issues</h3>
                <div class="value" style="color: #51cf66;">{len(mttr_all_resolved)}</div>
                <div class="label">已解決問題數</div>
                <div class="sub-stats">
                    <div class="sub-stat">
                        <div class="label">內部</div>
                        <div class="value" onclick="openMttrFilterInJira('resolved', 'internal')" style="cursor: pointer;">{len(mttr_resolved_internal)}</div>
                    </div>
                    <div class="sub-stat">
                        <div class="label">Vendor</div>
                        <div class="value" onclick="openMttrFilterInJira('resolved', 'vendor')" style="cursor: pointer;">{len(mttr_resolved_vendor)}</div>
                    </div>
                </div>
            </div>
            <div class="stat-card">
                <h3>Open Issues</h3>
                <div class="value" style="color: #ff6b6b;">{len(mttr_all_open)}</div>
                <div class="label">未解決問題數</div>
                <div class="sub-stats">
                    <div class="sub-stat">
                        <div class="label">內部</div>
                        <div class="value" onclick="openMttrFilterInJira('open', 'internal')" style="cursor: pointer;">{len(mttr_open_internal)}</div>
                    </div>
                    <div class="sub-stat">
                        <div class="label">Vendor</div>
                        <div class="value" onclick="openMttrFilterInJira('open', 'vendor')" style="cursor: pointer;">{len(mttr_open_vendor)}</div>
                    </div>
                </div>
            </div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR 指標說明</h2>
            <div style="background: #f8f9fa; padding: 15px; border-radius: 10px;">
                <p><strong>已解決問題 (Resolved):</strong></p>
                <ul style="margin: 10px 0 15px 20px;">
                    <li><strong>MTTR 指標:</strong> Resolved Date - Created Date (平均解決天數)</li>
                    <li><strong>平均 Overdue 天數:</strong> Resolved Date - Due Date (僅計算有 Due Date 且超期的問題)</li>
                </ul>
                <p><strong>未解決問題 (Open/Ongoing):</strong></p>
                <ul style="margin: 10px 0 0 20px;">
                    <li><strong>MTTR(ongoing) 指標:</strong> Now - Created Date (目前等待天數)</li>
                    <li><strong>平均 Overdue 天數:</strong> Now - Due Date (僅計算有 Due Date 且超期的問題)</li>
                </ul>
            </div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR 趨勢 - 已解決問題（內部 + Vendor）</h2>
            <div class="chart-wrapper"><canvas id="mttrResolvedAllChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR 趨勢 - 已解決問題 <span class="badge badge-internal">內部 JIRA</span></h2>
            <div class="chart-wrapper"><canvas id="mttrResolvedInternalChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR 趨勢 - 已解決問題 <span class="badge badge-vendor">Vendor JIRA</span></h2>
            <div class="chart-wrapper"><canvas id="mttrResolvedVendorChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR(ongoing) 趨勢 - 未解決問題（內部 + Vendor）</h2>
            <div class="chart-wrapper"><canvas id="mttrOpenAllChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR(ongoing) 趨勢 - 未解決問題 <span class="badge badge-internal">內部 JIRA</span></h2>
            <div class="chart-wrapper"><canvas id="mttrOpenInternalChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR(ongoing) 趨勢 - 未解決問題 <span class="badge badge-vendor">Vendor JIRA</span></h2>
            <div class="chart-wrapper"><canvas id="mttrOpenVendorChart"></canvas></div>
        </div>
"""

                # MTTR JavaScript 區塊
                mttr_js_section = f"""
        // ===== MTTR 圖表 =====
        const mttrFilterIds = {{
            resolved: {{
                internal: '{MTTR_FILTERS["resolved"]["internal"] or ""}',
                vendor: '{MTTR_FILTERS["resolved"]["vendor"] or ""}'
            }},
            open: {{
                internal: '{MTTR_FILTERS["open"]["internal"] or ""}',
                vendor: '{MTTR_FILTERS["open"]["vendor"] or ""}'
            }}
        }};

        const mttrWeekDates = {{
            resolved_internal: {mttr_resolved_internal_dates},
            resolved_vendor: {mttr_resolved_vendor_dates},
            open_internal: {mttr_open_internal_dates},
            open_vendor: {mttr_open_vendor_dates}
        }};

        function openMttrFilterInJira(type, source) {{
            const site = jiraSites[source];
            const filterId = mttrFilterIds[type][source];
            if (!filterId) return;

            let jql = `filter=${{filterId}}`;

            if (currentFilters.start_date) {{
                jql += ` AND created >= "${{currentFilters.start_date}} 00:00"`;
            }}
            if (currentFilters.end_date) {{
                jql += ` AND created <= "${{currentFilters.end_date}} 23:59"`;
            }}
            if (currentFilters.owner) {{
                jql += ` AND assignee="${{currentFilters.owner}}"`;
            }}

            console.log(`�� 跳轉 MTTR JIRA: ${{type}} (${{source}})`);
            console.log(`   JQL: ${{jql}}`);

            const url = `https://${{site}}/issues/?jql=${{encodeURIComponent(jql)}}`;
            window.open(url, '_blank');
        }}

        function openMttrWeekInJira(week, source, type) {{
            const site = jiraSites[source];
            const filterId = mttrFilterIds[type][source];
            if (!filterId) return;

            const dateKey = type + '_' + source;
            const weekData = mttrWeekDates[dateKey][week];
            if (!weekData) return;

            let jql = `filter=${{filterId}} AND created >= "${{weekData.start_date}} 00:00" AND created <= "${{weekData.end_date}} 23:59"`;
            const url = `https://${{site}}/issues/?jql=${{encodeURIComponent(jql)}}`;
            window.open(url, '_blank');
        }}

        function drawMttrChart(canvasId, labels, mttrData, overdueData, countData, source, type) {{
            const ctx = document.getElementById(canvasId).getContext('2d');
            const clickable = source !== null && type !== null;

            new Chart(ctx, {{
                type: 'line',
                data: {{
                    labels: labels,
                    datasets: [
                        {{
                            label: type === 'resolved' || type === null ? '平均 MTTR 天數' : '平均 MTTR(ongoing) 天數',
                            data: mttrData,
                            borderColor: '#667eea',
                            backgroundColor: 'rgba(102, 126, 234, 0.1)',
                            borderWidth: 3,
                            fill: true,
                            tension: 0.4,
                            yAxisID: 'y'
                        }},
                        {{
                            label: '平均 Overdue 天數',
                            data: overdueData,
                            borderColor: '#ff6b6b',
                            borderWidth: 3,
                            fill: false,
                            tension: 0.4,
                            yAxisID: 'y'
                        }},
                        {{
                            label: 'Issue 數量',
                            data: countData,
                            borderColor: '#51cf66',
                            borderWidth: 2,
                            borderDash: [5, 5],
                            fill: false,
                            tension: 0.4,
                            yAxisID: 'y1'
                        }}
                    ]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    interaction: {{ mode: 'index', intersect: false }},
                    onClick: clickable ? (event, elements) => {{
                        if (elements.length > 0) {{
                            const index = elements[0].index;
                            const week = labels[index];
                            openMttrWeekInJira(week, source, type);
                        }}
                    }} : undefined,
                    plugins: {{
                        legend: {{ display: true, position: 'top' }},
                        tooltip: {{
                            callbacks: {{
                                label: function(context) {{
                                    let label = context.dataset.label || '';
                                    if (label) label += ': ';
                                    if (context.parsed.y !== null) {{
                                        label += context.datasetIndex === 2 ? context.parsed.y + ' issues' : context.parsed.y.toFixed(2) + ' 天';
                                    }}
                                    return label;
                                }},
                                footer: clickable ? () => ['點擊查看 JIRA Issues'] : undefined
                            }}
                        }}
                    }},
                    scales: {{
                        y: {{
                            type: 'linear',
                            display: true,
                            position: 'left',
                            beginAtZero: true,
                            title: {{ display: true, text: '天數', color: '#667eea' }},
                            ticks: {{ color: '#667eea' }}
                        }},
                        y1: {{
                            type: 'linear',
                            display: true,
                            position: 'right',
                            beginAtZero: true,
                            title: {{ display: true, text: 'Issue 數量', color: '#51cf66' }},
                            ticks: {{ color: '#51cf66' }},
                            grid: {{ drawOnChartArea: false }}
                        }}
                    }}
                }}
            }});
        }}

        // 繪製 MTTR 圖表
        drawMttrChart('mttrResolvedAllChart', {mttr_resolved_all_labels}, {mttr_resolved_all_mttr}, {mttr_resolved_all_overdue}, {mttr_resolved_all_count}, null, 'resolved');
        drawMttrChart('mttrResolvedInternalChart', {mttr_resolved_internal_labels}, {mttr_resolved_internal_mttr}, {mttr_resolved_internal_overdue}, {mttr_resolved_internal_count}, 'internal', 'resolved');
        drawMttrChart('mttrResolvedVendorChart', {mttr_resolved_vendor_labels}, {mttr_resolved_vendor_mttr}, {mttr_resolved_vendor_overdue}, {mttr_resolved_vendor_count}, 'vendor', 'resolved');
        drawMttrChart('mttrOpenAllChart', {mttr_open_all_labels}, {mttr_open_all_mttr}, {mttr_open_all_overdue}, {mttr_open_all_count}, null, 'open');
        drawMttrChart('mttrOpenInternalChart', {mttr_open_internal_labels}, {mttr_open_internal_mttr}, {mttr_open_internal_overdue}, {mttr_open_internal_count}, 'internal', 'open');
        drawMttrChart('mttrOpenVendorChart', {mttr_open_vendor_labels}, {mttr_open_vendor_mttr}, {mttr_open_vendor_overdue}, {mttr_open_vendor_count}, 'vendor', 'open');

        console.log('✅ MTTR 圖表已載入');
"""
        
        # 過濾資料 - degrade 使用 created，resolved 使用 created
        filtered_degrade = filter_issues(data['degrade'], start_date, end_date, owner, date_field='created')
        filtered_resolved = filter_issues(data['resolved'], start_date, end_date, owner, date_field='created')
        
        # 修復 _source 標記
        for issue in filtered_degrade:
            if issue.get('_source') not in ['internal', 'vendor']:
                if 'vendorjira' in issue.get('self', '').lower():
                    issue['_source'] = 'vendor'
                else:
                    issue['_source'] = 'internal'
        
        for issue in filtered_resolved:
            if issue.get('_source') not in ['internal', 'vendor']:
                if 'vendorjira' in issue.get('self', '').lower():
                    issue['_source'] = 'vendor'
                else:
                    issue['_source'] = 'internal'
        
        # 分離內部和 Vendor
        internal_degrade = [i for i in filtered_degrade if i.get('_source') == 'internal']
        vendor_degrade = [i for i in filtered_degrade if i.get('_source') == 'vendor']
        internal_resolved = [i for i in filtered_resolved if i.get('_source') == 'internal']
        vendor_resolved = [i for i in filtered_resolved if i.get('_source') == 'vendor']
        
        manager = JiraDegradeManagerFast(
            site=JIRA_CONFIG['internal']['site'],
            user=JIRA_CONFIG['internal']['user'],
            password=JIRA_CONFIG['internal']['password'],
            token=JIRA_CONFIG['internal']['token']
        )
        
        # 統計分析 - degrade 使用 created，resolved 使用 created
        total_degrade = len(filtered_degrade)
        total_resolved = len(filtered_resolved)
        overall_percentage = (total_degrade / total_resolved * 100) if total_resolved > 0 else 0
        
        degrade_weekly = analyze_by_week_with_dates(filtered_degrade, date_field='created')
        resolved_weekly = analyze_by_week_with_dates(filtered_resolved, date_field='created')
        weekly_stats = calculate_weekly_percentage(degrade_weekly, resolved_weekly)
        
        degrade_weekly_internal = analyze_by_week_with_dates(internal_degrade, date_field='created')
        degrade_weekly_vendor = analyze_by_week_with_dates(vendor_degrade, date_field='created')
        resolved_weekly_internal = analyze_by_week_with_dates(internal_resolved, date_field='created')
        resolved_weekly_vendor = analyze_by_week_with_dates(vendor_resolved, date_field='created')
        
        degrade_assignees_internal = manager.get_assignee_distribution(internal_degrade)
        degrade_assignees_vendor = manager.get_assignee_distribution(vendor_degrade)
        resolved_assignees_internal = manager.get_assignee_distribution(internal_resolved)
        resolved_assignees_vendor = manager.get_assignee_distribution(vendor_resolved)
        
        # 週次趨勢數據（全部週次，不限制）
        trend_labels = json.dumps([w['week'] for w in weekly_stats])
        trend_data = json.dumps([w['percentage'] for w in weekly_stats])
        
        # 週次數量對比數據 - 新增 resolved_count
        count_degrade = json.dumps([w['degrade_count'] for w in weekly_stats])
        count_resolved = json.dumps([w['resolved_count'] for w in weekly_stats])
        
        # 週次分布數據（內部/Vendor）- 全部週次
        all_weeks_internal = sorted(set(list(degrade_weekly_internal.keys()) + list(resolved_weekly_internal.keys())))
        all_weeks_vendor = sorted(set(list(degrade_weekly_vendor.keys()) + list(resolved_weekly_vendor.keys())))
        
        weekly_internal_labels = json.dumps(all_weeks_internal)
        weekly_internal_degrade = json.dumps([degrade_weekly_internal.get(w, {}).get('count', 0) for w in all_weeks_internal])
        weekly_internal_resolved = json.dumps([resolved_weekly_internal.get(w, {}).get('count', 0) for w in all_weeks_internal])
        
        weekly_vendor_labels = json.dumps(all_weeks_vendor)
        weekly_vendor_degrade = json.dumps([degrade_weekly_vendor.get(w, {}).get('count', 0) for w in all_weeks_vendor])
        weekly_vendor_resolved = json.dumps([resolved_weekly_vendor.get(w, {}).get('count', 0) for w in all_weeks_vendor])
        
        # 依據 chart_limit 限制 Assignee 數據
        degrade_assignees_internal_top = dict(sorted(degrade_assignees_internal.items(), key=lambda x: x[1], reverse=True)[:chart_limit])
        degrade_assignees_vendor_top = dict(sorted(degrade_assignees_vendor.items(), key=lambda x: x[1], reverse=True)[:chart_limit])
        resolved_assignees_internal_top = dict(sorted(resolved_assignees_internal.items(), key=lambda x: x[1], reverse=True)[:chart_limit])
        resolved_assignees_vendor_top = dict(sorted(resolved_assignees_vendor.items(), key=lambda x: x[1], reverse=True)[:chart_limit])
        
        degrade_int_labels = json.dumps(list(degrade_assignees_internal_top.keys()))
        degrade_int_data = json.dumps(list(degrade_assignees_internal_top.values()))
        degrade_vnd_labels = json.dumps(list(degrade_assignees_vendor_top.keys()))
        degrade_vnd_data = json.dumps(list(degrade_assignees_vendor_top.values()))
        resolved_int_labels = json.dumps(list(resolved_assignees_internal_top.keys()))
        resolved_int_data = json.dumps(list(resolved_assignees_internal_top.values()))
        resolved_vnd_labels = json.dumps(list(resolved_assignees_vendor_top.keys()))
        resolved_vnd_data = json.dumps(list(resolved_assignees_vendor_top.values()))
        
        # 準備週次日期範圍數據（用於 JIRA 跳轉）
        weekly_date_ranges_degrade_internal = {}
        for week, stats in degrade_weekly_internal.items():
            weekly_date_ranges_degrade_internal[week] = {
                'start_date': stats.get('start_date'),
                'end_date': stats.get('end_date'),
                'start_datetime': stats.get('start_datetime'),
                'end_datetime': stats.get('end_datetime')
            }
        
        weekly_date_ranges_degrade_vendor = {}
        for week, stats in degrade_weekly_vendor.items():
            weekly_date_ranges_degrade_vendor[week] = {
                'start_date': stats.get('start_date'),
                'end_date': stats.get('end_date'),
                'start_datetime': stats.get('start_datetime'),
                'end_datetime': stats.get('end_datetime')
            }
        
        weekly_date_ranges_resolved_internal = {}
        for week, stats in resolved_weekly_internal.items():
            weekly_date_ranges_resolved_internal[week] = {
                'start_date': stats.get('start_date'),
                'end_date': stats.get('end_date'),
                'start_datetime': stats.get('start_datetime'),
                'end_datetime': stats.get('end_datetime')
            }
        
        weekly_date_ranges_resolved_vendor = {}
        for week, stats in resolved_weekly_vendor.items():
            weekly_date_ranges_resolved_vendor[week] = {
                'start_date': stats.get('start_date'),
                'end_date': stats.get('end_date'),
                'start_datetime': stats.get('start_datetime'),
                'end_datetime': stats.get('end_datetime')
            }
        
        # 轉換為 JSON
        date_ranges_degrade_internal_json = json.dumps(weekly_date_ranges_degrade_internal)
        date_ranges_degrade_vendor_json = json.dumps(weekly_date_ranges_degrade_vendor)
        date_ranges_resolved_internal_json = json.dumps(weekly_date_ranges_resolved_internal)
        date_ranges_resolved_vendor_json = json.dumps(weekly_date_ranges_resolved_vendor)
        
        # JIRA sites 和 filter IDs
        jira_sites_json = json.dumps(data['jira_sites'])
        filter_ids_json = json.dumps(FILTERS)
        
        # 當前過濾條件
        current_filters_json = json.dumps({
            'start_date': start_date or '',
            'end_date': end_date or '',
            'owner': owner or ''
        })
        
        # 準備表格數據
        def generate_assignee_table_html(assignee_dict, source, type_name, chart_limit):
            """生成 Assignee 表格 HTML"""
            sorted_data = sorted(assignee_dict.items(), key=lambda x: x[1], reverse=True)[:chart_limit]
            total = sum(assignee_dict.values())
            
            site = data['jira_sites'][source]
            filter_id = FILTERS[type_name][source]
            
            # 根據 type 使用不同的日期欄位
            date_field = 'created' if type_name == 'degrade' else 'created'
            
            html = '<table style="width: 100%; border-collapse: collapse;">'
            html += '<thead><tr style="background: #667eea; color: white;">'
            html += '<th style="padding: 12px; text-align: left;">排名</th>'
            html += '<th style="padding: 12px; text-align: left;">Assignee</th>'
            html += '<th style="padding: 12px; text-align: left;">Count</th>'
            html += '<th style="padding: 12px; text-align: left;">Percentage</th>'
            html += '</tr></thead><tbody>'
            
            for index, (name, count) in enumerate(sorted_data, 1):
                percentage = (count / total * 100) if total > 0 else 0
                
                # 建立 JIRA 連結 - 使用對應的日期欄位
                jql = f'filter={filter_id} AND assignee="{name}"'
                if start_date:
                    jql += f' AND {date_field} >= "{start_date} 00:00"'
                if end_date:
                    jql += f' AND {date_field} <= "{end_date} 23:59"'
                
                url = f"https://{site}/issues/?jql={quote(jql)}"
                
                bg_color = '#f5f5f5' if index % 2 == 0 else 'white'
                html += f'<tr style="background: {bg_color};">'
                html += f'<td style="padding: 12px; border-bottom: 1px solid #eee;">{index}</td>'
                html += f'<td style="padding: 12px; border-bottom: 1px solid #eee;"><a href="{url}" target="_blank" style="color: #667eea; text-decoration: none; font-weight: 500;">{name}</a></td>'
                html += f'<td style="padding: 12px; border-bottom: 1px solid #eee;">{count}</td>'
                html += f'<td style="padding: 12px; border-bottom: 1px solid #eee;">{percentage:.2f}%</td>'
                html += '</tr>'
            
            html += '</tbody></table>'
            return html
        
        table_degrade_internal = generate_assignee_table_html(degrade_assignees_internal, 'internal', 'degrade', chart_limit)
        table_degrade_vendor = generate_assignee_table_html(degrade_assignees_vendor, 'vendor', 'degrade', chart_limit)
        table_resolved_internal = generate_assignee_table_html(resolved_assignees_internal, 'internal', 'resolved', chart_limit)
        table_resolved_vendor = generate_assignee_table_html(resolved_assignees_vendor, 'vendor', 'resolved', chart_limit)
        
        # 生成 HTML
        html_content = f"""
<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>JIRA Degrade % 分析報告</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1600px;
            margin: 0 auto;
        }}
        
        .header {{
            background: white;
            padding: 30px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            margin-bottom: 30px;
        }}
        
        .header h1 {{
            color: #333;
            font-size: 2.2em;
            margin-bottom: 10px;
        }}
        
        .header p {{
            color: #666;
            font-size: 1em;
        }}
        
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .stat-card {{
            background: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            text-align: center;
            transition: transform 0.3s;
        }}
        
        .stat-card:hover {{
            transform: translateY(-5px);
        }}
        
        .stat-card h3 {{
            color: #666;
            font-size: 1em;
            margin-bottom: 10px;
            text-transform: uppercase;
        }}
        
        .stat-card .value {{
            font-size: 3em;
            font-weight: bold;
            color: #667eea;
        }}
        
        .stat-card .sub-stats {{
            display: flex;
            justify-content: space-around;
            margin-top: 15px;
            padding-top: 15px;
            border-top: 1px solid #eee;
        }}
        
        .stat-card .sub-stat .label {{
            font-size: 0.8em;
            color: #999;
        }}
        
        .stat-card .sub-stat .value {{
            font-size: 1.2em;
            font-weight: bold;
            color: #667eea;
        }}
        
        .chart-container {{
            background: white;
            padding: 30px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            margin-bottom: 30px;
        }}
        
        .chart-container h2 {{
            color: #333;
            margin-bottom: 20px;
            font-size: 1.5em;
        }}
        
        .chart-wrapper {{
            position: relative;
            height: 400px;
            cursor: pointer;
        }}
        
        .chart-wrapper-dynamic {{
            position: relative;
            min-height: 400px;
            cursor: pointer;
        }}
        
        .badge {{
            display: inline-block;
            padding: 3px 10px;
            border-radius: 12px;
            font-size: 0.7em;
            margin-left: 10px;
        }}
        
        .badge-internal {{
            background: #e3f2fd;
            color: #1976d2;
        }}
        
        .badge-vendor {{
            background: #fff3e0;
            color: #f57c00;
        }}
        
        .info-banner {{
            background: #e3f2fd;
            color: #1976d2;
            padding: 15px 20px;
            border-radius: 10px;
            margin-bottom: 20px;
            text-align: center;
            font-size: 0.95em;
        }}
        
        .info-banner strong {{
            font-weight: 600;
        }}
        
        .table-container {{
            background: white;
            padding: 30px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            margin-bottom: 30px;
        }}
        
        .table-container h2 {{
            color: #333;
            margin-bottom: 20px;
            font-size: 1.3em;
        }}

        /* 頁籤樣式 */
        .tab-navigation {{
            display: flex;
            gap: 10px;
            margin-bottom: 30px;
        }}

        .tab-button {{
            padding: 15px 30px;
            border: 2px solid #999;
            border-radius: 10px 10px 0 0;
            background: linear-gradient(180deg, #ffffff 0%, #d0d0d0 100%);
            color: #333;
            font-size: 1.1em;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
        }}

        .tab-button:hover {{
            background: linear-gradient(180deg, #f0f0f0 0%, #c0c0c0 100%);
        }}

        .tab-button.active {{
            background: linear-gradient(180deg, #7b8ff5 0%, #667eea 100%);
            color: white;
            border-color: #667eea;
        }}

        .tab-content {{
            display: none;
        }}

        .tab-content.active {{
            display: block;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1 id="pageTitle">�� JIRA Degrade % 分析報告</h1>
            <p id="pageDesc">公版 SQA/QC Degrade 問題統計分析</p>
            <p style="margin-top: 10px; font-size: 0.9em; color: #999;">
                生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 圖表顯示筆數: {chart_limit}
            </p>
            <p style="margin-top: 5px; font-size: 0.85em; color: #999;">
                �� Degrade 使用 created 日期 | Resolved 使用 created 日期
            </p>
        </div>
        
        <!-- 頁籤導航 -->
        <div class="tab-navigation">
            <button class="tab-button active" onclick="switchTab('degrade')">�� Degrade 分析</button>
            {'<button class="tab-button" onclick="switchTab(\'mttr\')">�� MTTR 指標</button>' if MTTR_ENABLED else ''}
        </div>

        <!-- Degrade 頁籤內容 -->
        <div id="degradeTab" class="tab-content active">

        <div class="info-banner">
            <strong>�� 提示：</strong> 圖表可以點擊！點擊週次 bar 可跳轉到 JIRA 查看該週的 issues，點擊 Assignee bar 可查看該人員的所有 issues
        </div>
        
        <div class="stats-grid">
            <div class="stat-card">
                <h3>Degrade Issues</h3>
                <div class="value" id="degradeCount">{total_degrade}</div>
                <div class="label">問題總數</div>
                <div class="sub-stats">
                    <div class="sub-stat">
                        <div class="label">內部</div>
                        <div class="value" onclick="openFilterInJira('degrade', 'internal')" style="cursor: pointer;">{len(internal_degrade)}</div>
                    </div>
                    <div class="sub-stat">
                        <div class="label">Vendor</div>
                        <div class="value" onclick="openFilterInJira('degrade', 'vendor')" style="cursor: pointer;">{len(vendor_degrade)}</div>
                    </div>
                </div>
            </div>
            <div class="stat-card">
                <h3>Resolved Issues</h3>
                <div class="value" id="resolvedCount">{total_resolved}</div>
                <div class="label">解題總數</div>
                <div class="sub-stats">
                    <div class="sub-stat">
                        <div class="label">內部</div>
                        <div class="value" onclick="openFilterInJira('resolved', 'internal')" style="cursor: pointer;">{len(internal_resolved)}</div>
                    </div>
                    <div class="sub-stat">
                        <div class="label">Vendor</div>
                        <div class="value" onclick="openFilterInJira('resolved', 'vendor')" style="cursor: pointer;">{len(vendor_resolved)}</div>
                    </div>
                </div>
            </div>
            <div class="stat-card">
                <h3>Degrade %</h3>
                <div class="value">{overall_percentage:.2f}%</div>
                <div class="label">整體比例</div>
            </div>
        </div>
        
        <!-- 每週趨勢圖 -->
        <div class="chart-container">
            <h2>�� 每週 Degrade % 與 軸：Degrade & CCC issue 數量趨勢（內部 + Vendor 合併）</h2>
            <p style="color: #666; font-size: 0.9em; margin-top: 5px;">
                �� 左側 Y 軸：Degrade % | 右側 Y 軸：軸：Degrade & CCC issue 數量
                <br>�� Degrade 使用 created 日期 | Resolved 使用 created 日期
            </p>                
            <div class="chart-wrapper">
                <canvas id="trendChart"></canvas>
            </div>
        </div>
        
        <div class="chart-container">
            <h2>�� 每週 Degrade vs CCC issue 數量</h2>
            <div class="chart-wrapper">
                <canvas id="countChart"></canvas>
            </div>
        </div>
        
        <div class="chart-container">
            <h2>�� 每週 Degrade 數量分布 <span class="badge badge-internal">內部 JIRA</span> <small style="color: #999;">（點擊可跳轉 JIRA）</small></h2>
            <div class="chart-wrapper">
                <canvas id="weeklyDegradeInternal"></canvas>
            </div>
        </div>
        
        <div class="chart-container">
            <h2>�� 每週 Degrade 數量分布 <span class="badge badge-vendor">Vendor JIRA</span> <small style="color: #999;">（點擊可跳轉 JIRA）</small></h2>
            <div class="chart-wrapper">
                <canvas id="weeklyDegradeVendor"></canvas>
            </div>
        </div>
        
        <div class="chart-container">
            <h2>�� Degrade Issues Assignee 分布 <span class="badge badge-internal">內部 JIRA</span> <small style="color: #999;">（Top {chart_limit}，點擊可跳轉 JIRA）</small></h2>
            <div class="chart-wrapper-dynamic" id="degradeAssigneeInternalWrapper">
                <canvas id="degradeAssigneeInternal"></canvas>
            </div>
        </div>
        
        <div class="chart-container">
            <h2>�� Degrade Issues Assignee 分布 <span class="badge badge-vendor">Vendor JIRA</span> <small style="color: #999;">（Top {chart_limit}，點擊可跳轉 JIRA）</small></h2>
            <div class="chart-wrapper-dynamic" id="degradeAssigneeVendorWrapper">
                <canvas id="degradeAssigneeVendor"></canvas>
            </div>
        </div>
        
        <div class="chart-container">
            <h2>�� Resolved Issues Assignee 分布 <span class="badge badge-internal">內部 JIRA</span> <small style="color: #999;">（Top {chart_limit}，點擊可跳轉 JIRA）</small></h2>
            <div class="chart-wrapper-dynamic" id="resolvedAssigneeInternalWrapper">
                <canvas id="resolvedAssigneeInternal"></canvas>
            </div>
        </div>
        
        <div class="chart-container">
            <h2>�� Resolved Issues Assignee 分布 <span class="badge badge-vendor">Vendor JIRA</span> <small style="color: #999;">（Top {chart_limit}，點擊可跳轉 JIRA）</small></h2>
            <div class="chart-wrapper-dynamic" id="resolvedAssigneeVendorWrapper">
                <canvas id="resolvedAssigneeVendor"></canvas>
            </div>
        </div>
        
        <div class="table-container">
            <h2>�� Degrade Issues Assignee 詳細分布 <span class="badge badge-internal">內部 JIRA</span> <small style="color: #999;">（Top {chart_limit}）</small></h2>
            {table_degrade_internal}
        </div>
        
        <div class="table-container">
            <h2>�� Degrade Issues Assignee 詳細分布 <span class="badge badge-vendor">Vendor JIRA</span> <small style="color: #999;">（Top {chart_limit}）</small></h2>
            {table_degrade_vendor}
        </div>
        
        <div class="table-container">
            <h2>�� Resolved Issues Assignee 詳細分布 <span class="badge badge-internal">內部 JIRA</span> <small style="color: #999;">（Top {chart_limit}）</small></h2>
            {table_resolved_internal}
        </div>
        
        <div class="table-container">
            <h2>�� Resolved Issues Assignee 詳細分布 <span class="badge badge-vendor">Vendor JIRA</span> <small style="color: #999;">（Top {chart_limit}）</small></h2>
            {table_resolved_vendor}
        </div>

        </div><!-- 結束 Degrade 頁籤 -->

        <!-- MTTR 頁籤內容 -->
        <div id="mttrTab" class="tab-content">
        {mttr_html_section}
        </div><!-- 結束 MTTR 頁籤 -->

        <footer style="background: white; padding: 20px; border-radius: 15px; box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2); margin-top: 30px; text-align: center; color: #666;">
            <p style="margin: 0; font-weight: 500;">© 2025 Copyright by Vince. All rights reserved.</p>
            <p style="margin-top: 8px; font-size: 0.85em; color: #999;">CCC Degrade % 分析報告{' + MTTR 指標' if MTTR_ENABLED else ''}</p>
        </footer>
    </div>
    
    <script>
        const jiraSites = {jira_sites_json};
        const filterIds = {filter_ids_json};
        const currentFilters = {current_filters_json};
        
        const weeklyDateRanges = {{
            degrade_internal: {date_ranges_degrade_internal_json},
            degrade_vendor: {date_ranges_degrade_vendor_json},
            resolved_internal: {date_ranges_resolved_internal_json},
            resolved_vendor: {date_ranges_resolved_vendor_json}
        }};

        // 頁籤切換函數
        function switchTab(tabName) {{
            // 隱藏所有頁籤內容
            document.querySelectorAll('.tab-content').forEach(content => {{
                content.classList.remove('active');
            }});

            // 移除所有頁籤按鈕的 active 狀態
            document.querySelectorAll('.tab-button').forEach(button => {{
                button.classList.remove('active');
            }});

            // 顯示選中的頁籤內容
            const tabContent = document.getElementById(tabName + 'Tab');
            if (tabContent) {{
                tabContent.classList.add('active');
            }}

            // 設置選中的頁籤按鈕為 active
            event.target.classList.add('active');

            // 更新標題
            const pageTitle = document.getElementById('pageTitle');
            const pageDesc = document.getElementById('pageDesc');

            if (tabName === 'degrade') {{
                pageTitle.innerHTML = '�� JIRA Degrade % 分析報告';
                pageDesc.textContent = '公版 SQA/QC Degrade 問題統計分析';
                document.title = 'JIRA Degrade % 分析報告';
            }} else if (tabName === 'mttr') {{
                pageTitle.innerHTML = '�� MTTR 指標分析報告';
                pageDesc.textContent = 'Mean Time To Resolve - 平均解決時間分析';
                document.title = 'MTTR 指標分析報告';
            }}
        }}

        function openFilterInJira(type, source) {{
            const site = source === 'internal' ? jiraSites.internal : jiraSites.vendor;
            const filterId = filterIds[type][source];
            
            let dateField = type === 'degrade' ? 'created' : 'created';
            let jql = `filter=${{filterId}}`;
            
            if (currentFilters.start_date) {{
                jql += ` AND ${{dateField}} >= "${{currentFilters.start_date}} 00:00"`;
            }}
            if (currentFilters.end_date) {{
                jql += ` AND ${{dateField}} <= "${{currentFilters.end_date}} 23:59"`;
            }}
            if (currentFilters.owner) {{
                jql += ` AND assignee="${{currentFilters.owner}}"`;
            }}
            
            console.log(`�� 跳轉 JIRA: ${{type}} (${{source}})`);
            console.log(`   JQL: ${{jql}}`);
            
            const url = `https://${{site}}/issues/?jql=${{encodeURIComponent(jql)}}`;
            window.open(url, '_blank');
        }}

        function openWeekInJira(week, source, type) {{
            const site = source === 'internal' ? jiraSites.internal : jiraSites.vendor;
            const filterId = filterIds[type][source];
            
            const dateRangesKey = `${{type}}_${{source}}`;
            const dateRanges = weeklyDateRanges[dateRangesKey];
            
            if (!dateRanges || !dateRanges[week]) {{
                alert(`無法找到週次 ${{week}} 的日期範圍`);
                return;
            }}
            
            const weekStartDate = dateRanges[week].start_date;
            const weekEndDate = dateRanges[week].end_date;
            
            // 根據 type 使用不同的日期欄位
            const dateField = type === 'degrade' ? 'created' : 'created';
            let jql = `filter=${{filterId}} AND ${{dateField}} >= "${{weekStartDate}} 00:00" AND ${{dateField}} <= "${{weekEndDate}} 23:59"`;
            
            if (currentFilters.owner) {{
                jql += ` AND assignee="${{currentFilters.owner}}"`;
            }}
            
            console.log(`�� 跳轉 JIRA: 週次 ${{week}} (${{source}}, ${{type}})`);
            console.log(`   JQL: ${{jql}}`);
            
            const url = `https://${{site}}/issues/?jql=${{encodeURIComponent(jql)}}`;
            window.open(url, '_blank');
        }}
        
        function openAssigneeInJira(assigneeName, source, type) {{
            const site = source === 'internal' ? jiraSites.internal : jiraSites.vendor;
            const filterId = filterIds[type][source];
            
            // 根據 type 使用不同的日期欄位
            const dateField = type === 'degrade' ? 'created' : 'created';
            let jql = `filter=${{filterId}} AND assignee="${{assigneeName}}"`;
            
            if (currentFilters.start_date) {{
                jql += ` AND ${{dateField}} >= "${{currentFilters.start_date}} 00:00"`;
            }}
            if (currentFilters.end_date) {{
                jql += ` AND ${{dateField}} <= "${{currentFilters.end_date}} 23:59"`;
            }}
            
            console.log(`�� 跳轉 JIRA: Assignee ${{assigneeName}} (${{source}}, ${{type}})`);
            console.log(`   JQL: ${{jql}}`);
            
            const url = `https://${{site}}/issues/?jql=${{encodeURIComponent(jql)}}`;
            window.open(url, '_blank');
        }}
        
        // 趨勢圖 - 三條線（Degrade 數量 + CCC issue 數量 + Degrade % 參考線）
        new Chart(document.getElementById('trendChart'), {{
            type: 'line',
            data: {{
                labels: {trend_labels},
                datasets: [
                    {{
                        label: 'Degrade 數量',
                        data: {count_degrade},
                        borderColor: '#ff6b6b',
                        backgroundColor: 'rgba(255, 107, 107, 0.1)',
                        borderWidth: 3,
                        fill: true,
                        tension: 0.4,
                        pointRadius: 5,
                        pointHoverRadius: 7,
                        yAxisID: 'y1'
                    }},
                    {{
                        label: 'CCC issue 數量',
                        data: {count_resolved},
                        borderColor: '#51cf66',
                        backgroundColor: 'rgba(81, 207, 102, 0.1)',
                        borderWidth: 3,
                        fill: false,
                        tension: 0.4,
                        pointRadius: 5,
                        pointHoverRadius: 7,
                        yAxisID: 'y1'
                    }},
                    {{
                        label: 'Degrade % (參考)',
                        data: {trend_data},
                        borderColor: '#667eea',
                        backgroundColor: 'transparent',
                        borderWidth: 2,
                        borderDash: [5, 5],
                        fill: false,
                        tension: 0.4,
                        pointRadius: 3,
                        pointHoverRadius: 5,
                        yAxisID: 'y2',
                        hidden: false
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                interaction: {{
                    mode: 'index',
                    intersect: false
                }},
                plugins: {{
                    legend: {{ 
                        display: true,
                        position: 'top',
                        labels: {{
                            usePointStyle: true,
                            padding: 15
                        }}
                    }},
                    tooltip: {{
                        callbacks: {{
                            label: function(context) {{
                                let label = context.dataset.label || '';
                                if (label) {{
                                    label += ': ';
                                }}
                                if (context.parsed.y !== null) {{
                                    if (context.datasetIndex === 2) {{
                                        label += context.parsed.y.toFixed(2) + '%';
                                    }} else {{
                                        label += context.parsed.y + ' issues';
                                    }}
                                }}
                                return label;
                            }}
                        }}
                    }}
                }},
                scales: {{
                    y: {{
                        type: 'linear',
                        display: false,
                        position: 'left',
                        beginAtZero: true,
                        title: {{ 
                            display: true, 
                            text: 'Degrade 數量',
                            color: '#ff6b6b',
                            font: {{
                                size: 14,
                                weight: 'bold'
                            }}
                        }},
                        ticks: {{
                            color: '#ff6b6b'
                        }}
                    }},
                    y1: {{
                        type: 'linear',
                        display: true,
                        position: 'right',
                        beginAtZero: true,
                        title: {{ 
                            display: true, 
                            text: 'CCC issue 數量',
                            color: '#51cf66',
                            font: {{
                                size: 14,
                                weight: 'bold'
                            }}
                        }},
                        ticks: {{
                            color: '#51cf66'
                        }},
                        grid: {{
                            drawOnChartArea: false
                        }}
                    }},
                    y2: {{
                        type: 'linear',
                        display: true,
                        position: 'left',
                        beginAtZero: true,
                        title: {{ 
                            display: true, 
                            text: 'Degrade %',
                            color: '#667eea',
                            font: {{
                                size: 14,
                                weight: 'bold'
                            }}
                        }},
                        ticks: {{
                            color: '#667eea'
                        }},
                        max: 100,
                        grid: {{
                            drawOnChartArea: false
                        }}
                    }}
                }}
            }}
        }});
        
        // 數量圖
        new Chart(document.getElementById('countChart'), {{
            type: 'bar',
            data: {{
                labels: {trend_labels},
                datasets: [
                    {{
                        label: 'Degrade Issues',
                        data: {count_degrade},
                        backgroundColor: '#ff6b6b'
                    }},
                    {{
                        label: 'Resolved Issues',
                        data: {count_resolved},
                        backgroundColor: '#51cf66'
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{ display: true }}
                }}
            }}
        }});
        
        // 週次 Degrade 分布 - 內部（可點擊）
        new Chart(document.getElementById('weeklyDegradeInternal'), {{
            type: 'bar',
            data: {{
                labels: {weekly_internal_labels},
                datasets: [
                    {{
                        label: 'Degrade Issues',
                        data: {weekly_internal_degrade},
                        backgroundColor: '#ff6b6b'
                    }},
                    {{
                        label: 'Resolved Issues',
                        data: {weekly_internal_resolved},
                        backgroundColor: '#51cf66'
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                onClick: (event, elements) => {{
                    if (elements.length > 0) {{
                        const index = elements[0].index;
                        const datasetIndex = elements[0].datasetIndex;
                        const weeks = {weekly_internal_labels};
                        const week = weeks[index];
                        const type = datasetIndex === 0 ? 'degrade' : 'resolved';
                        openWeekInJira(week, 'internal', type);
                    }}
                }},
                plugins: {{
                    legend: {{ display: true }},
                    tooltip: {{
                        callbacks: {{
                            afterBody: () => ['', '�� 點擊可跳轉到 JIRA 查看該週的 issues']
                        }}
                    }}
                }}
            }}
        }});
        
        // 週次 Degrade 分布 - Vendor（可點擊）
        new Chart(document.getElementById('weeklyDegradeVendor'), {{
            type: 'bar',
            data: {{
                labels: {weekly_vendor_labels},
                datasets: [
                    {{
                        label: 'Degrade Issues',
                        data: {weekly_vendor_degrade},
                        backgroundColor: '#ff6b6b'
                    }},
                    {{
                        label: 'Resolved Issues',
                        data: {weekly_vendor_resolved},
                        backgroundColor: '#51cf66'
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                onClick: (event, elements) => {{
                    if (elements.length > 0) {{
                        const index = elements[0].index;
                        const datasetIndex = elements[0].datasetIndex;
                        const weeks = {weekly_vendor_labels};
                        const week = weeks[index];
                        const type = datasetIndex === 0 ? 'degrade' : 'resolved';
                        openWeekInJira(week, 'vendor', type);
                    }}
                }},
                plugins: {{
                    legend: {{ display: true }},
                    tooltip: {{
                        callbacks: {{
                            afterBody: () => ['', '�� 點擊可跳轉到 JIRA 查看該週的 issues']
                        }}
                    }}
                }}
            }}
        }});
        
        // 動態高度 Assignee 圖表函數
        function drawAssigneeChart(canvasId, labels, data, chartLabel, color, source, type) {{
            const ctx = document.getElementById(canvasId).getContext('2d');
            const chartHeight = Math.max(400, labels.length * 30);
            const wrapper = document.getElementById(canvasId + 'Wrapper');
            if (wrapper) {{
                wrapper.style.height = chartHeight + 'px';
            }}
            
            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: labels,
                    datasets: [{{
                        label: chartLabel,
                        data: data,
                        backgroundColor: color
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    indexAxis: 'y',
                    onClick: (event, elements) => {{
                        if (elements.length > 0) {{
                            const index = elements[0].index;
                            const assigneeName = labels[index];
                            openAssigneeInJira(assigneeName, source, type);
                        }}
                    }},
                    plugins: {{
                        legend: {{ display: true }},
                        tooltip: {{
                            callbacks: {{
                                afterBody: () => ['', '�� 點擊可跳轉到 JIRA 查看該 Assignee 的 issues']
                            }}
                        }}
                    }}
                }}
            }});
        }}
        
        // Assignee 圖表
        drawAssigneeChart('degradeAssigneeInternal', {degrade_int_labels}, {degrade_int_data}, 'Degrade Issues', '#ff6b6b', 'internal', 'degrade');
        drawAssigneeChart('degradeAssigneeVendor', {degrade_vnd_labels}, {degrade_vnd_data}, 'Degrade Issues', '#ff6b6b', 'vendor', 'degrade');
        drawAssigneeChart('resolvedAssigneeInternal', {resolved_int_labels}, {resolved_int_data}, 'Resolved Issues', '#51cf66', 'internal', 'resolved');
        drawAssigneeChart('resolvedAssigneeVendor', {resolved_vnd_labels}, {resolved_vnd_data}, 'Resolved Issues', '#51cf66', 'vendor', 'resolved');
        
        console.log('✅ Degrade 圖表已載入');

        {mttr_js_section}
    </script>
</body>
</html>
"""
        
        output = io.BytesIO(html_content.encode('utf-8'))
        output.seek(0)
        
        filename = f"jira_degrade_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        return send_file(
            output,
            mimetype='text/html',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ HTML 匯出失敗: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/mttr/html')
def export_mttr_html():
    """匯出 MTTR HTML 報告"""
    if not MTTR_ENABLED:
        return jsonify({'success': False, 'error': 'MTTR 功能未啟用'}), 400

    try:
        data = get_mttr_data()
        if not data:
            return jsonify({'success': False, 'error': '無法載入 MTTR 資料'}), 500

        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner')

        # 過濾資料
        resolved_internal = filter_issues(data['resolved']['internal'], start_date, end_date, owner, date_field='created')
        resolved_vendor = filter_issues(data['resolved']['vendor'], start_date, end_date, owner, date_field='created')
        open_internal = filter_issues(data['open']['internal'], start_date, end_date, owner, date_field='created')
        open_vendor = filter_issues(data['open']['vendor'], start_date, end_date, owner, date_field='created')

        all_resolved = resolved_internal + resolved_vendor
        all_open = open_internal + open_vendor

        # 計算 MTTR 指標
        resolved_stats_all = calculate_mttr_metrics(all_resolved, 'resolved')
        resolved_stats_internal = calculate_mttr_metrics(resolved_internal, 'resolved')
        resolved_stats_vendor = calculate_mttr_metrics(resolved_vendor, 'resolved')

        open_stats_all = calculate_mttr_metrics(all_open, 'open')
        open_stats_internal = calculate_mttr_metrics(open_internal, 'open')
        open_stats_vendor = calculate_mttr_metrics(open_vendor, 'open')

        # 準備圖表數據
        resolved_all_labels = json.dumps([w['week'] for w in resolved_stats_all])
        resolved_all_mttr = json.dumps([w['avg_mttr_days'] for w in resolved_stats_all])
        resolved_all_overdue = json.dumps([w['avg_overdue_days'] for w in resolved_stats_all])
        resolved_all_count = json.dumps([w['count'] for w in resolved_stats_all])

        resolved_internal_labels = json.dumps([w['week'] for w in resolved_stats_internal])
        resolved_internal_mttr = json.dumps([w['avg_mttr_days'] for w in resolved_stats_internal])
        resolved_internal_overdue = json.dumps([w['avg_overdue_days'] for w in resolved_stats_internal])
        resolved_internal_count = json.dumps([w['count'] for w in resolved_stats_internal])

        resolved_vendor_labels = json.dumps([w['week'] for w in resolved_stats_vendor])
        resolved_vendor_mttr = json.dumps([w['avg_mttr_days'] for w in resolved_stats_vendor])
        resolved_vendor_overdue = json.dumps([w['avg_overdue_days'] for w in resolved_stats_vendor])
        resolved_vendor_count = json.dumps([w['count'] for w in resolved_stats_vendor])

        open_all_labels = json.dumps([w['week'] for w in open_stats_all])
        open_all_mttr = json.dumps([w['avg_mttr_days'] for w in open_stats_all])
        open_all_overdue = json.dumps([w['avg_overdue_days'] for w in open_stats_all])
        open_all_count = json.dumps([w['count'] for w in open_stats_all])

        open_internal_labels = json.dumps([w['week'] for w in open_stats_internal])
        open_internal_mttr = json.dumps([w['avg_mttr_days'] for w in open_stats_internal])
        open_internal_overdue = json.dumps([w['avg_overdue_days'] for w in open_stats_internal])
        open_internal_count = json.dumps([w['count'] for w in open_stats_internal])

        open_vendor_labels = json.dumps([w['week'] for w in open_stats_vendor])
        open_vendor_mttr = json.dumps([w['avg_mttr_days'] for w in open_stats_vendor])
        open_vendor_overdue = json.dumps([w['avg_overdue_days'] for w in open_stats_vendor])
        open_vendor_count = json.dumps([w['count'] for w in open_stats_vendor])

        # 準備週次日期範圍（用於 JIRA 連結）
        resolved_internal_dates = json.dumps({w['week']: {'start_date': w['start_date'], 'end_date': w['end_date']} for w in resolved_stats_internal})
        resolved_vendor_dates = json.dumps({w['week']: {'start_date': w['start_date'], 'end_date': w['end_date']} for w in resolved_stats_vendor})
        open_internal_dates = json.dumps({w['week']: {'start_date': w['start_date'], 'end_date': w['end_date']} for w in open_stats_internal})
        open_vendor_dates = json.dumps({w['week']: {'start_date': w['start_date'], 'end_date': w['end_date']} for w in open_stats_vendor})

        filter_info = f"篩選條件: {start_date or '不限'} ~ {end_date or '不限'}" + (f", Assignee: {owner}" if owner else "")

        html_content = f"""
<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>MTTR 指標分析報告</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        .container {{ max-width: 1600px; margin: 0 auto; }}
        .header {{
            background: white;
            padding: 30px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            margin-bottom: 30px;
        }}
        .header h1 {{ color: #333; font-size: 2.2em; margin-bottom: 10px; }}
        .header p {{ color: #666; font-size: 1em; }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .stat-card {{
            background: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            text-align: center;
        }}
        .stat-card h3 {{ color: #333; margin-bottom: 10px; }}
        .stat-card .value {{ font-size: 2.5em; font-weight: bold; color: #667eea; }}
        .stat-card .label {{ color: #666; font-size: 0.9em; }}
        .chart-container {{
            background: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            margin-bottom: 30px;
        }}
        .chart-container h2 {{ color: #333; margin-bottom: 20px; font-size: 1.3em; }}
        .chart-wrapper {{ height: 400px; position: relative; }}
        .badge {{
            display: inline-block;
            padding: 3px 10px;
            border-radius: 12px;
            font-size: 0.75em;
            font-weight: 500;
        }}
        .badge-internal {{ background: #e3f2fd; color: #1976d2; }}
        .badge-vendor {{ background: #fce4ec; color: #c2185b; }}
        .info-box {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 20px;
        }}
        .info-box p {{ margin: 5px 0; }}
        .info-box ul {{ margin: 10px 0 10px 20px; }}
        footer {{
            background: white;
            padding: 20px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            margin-top: 30px;
            text-align: center;
            color: #666;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>�� MTTR 指標分析報告</h1>
            <p>匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>{filter_info}</p>
        </div>

        <div class="stats-grid">
            <div class="stat-card">
                <h3>Resolved Issues</h3>
                <div class="value">{len(all_resolved)}</div>
                <div class="label">已解決問題數（內部: {len(resolved_internal)} / Vendor: {len(resolved_vendor)}）</div>
            </div>
            <div class="stat-card">
                <h3>Open Issues</h3>
                <div class="value">{len(all_open)}</div>
                <div class="label">未解決問題數（內部: {len(open_internal)} / Vendor: {len(open_vendor)}）</div>
            </div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR 指標說明</h2>
            <div class="info-box">
                <p><strong>已解決問題 (Resolved):</strong></p>
                <ul>
                    <li><strong>MTTR 指標:</strong> Resolved Date - Created Date (平均解決天數)</li>
                    <li><strong>平均 Overdue 天數:</strong> Resolved Date - Due Date (僅計算有 Due Date 且超期的問題)</li>
                </ul>
                <p><strong>未解決問題 (Open/Ongoing):</strong></p>
                <ul>
                    <li><strong>MTTR(ongoing) 指標:</strong> Now - Created Date (目前等待天數)</li>
                    <li><strong>平均 Overdue 天數:</strong> Now - Due Date (僅計算有 Due Date 且超期的問題)</li>
                </ul>
            </div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR 趨勢 - 已解決問題（內部 + Vendor）</h2>
            <div class="chart-wrapper"><canvas id="resolvedAllChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR 趨勢 - 已解決問題 <span class="badge badge-internal">內部 JIRA</span></h2>
            <div class="chart-wrapper"><canvas id="resolvedInternalChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR 趨勢 - 已解決問題 <span class="badge badge-vendor">Vendor JIRA</span></h2>
            <div class="chart-wrapper"><canvas id="resolvedVendorChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR(ongoing) 趨勢 - 未解決問題（內部 + Vendor）</h2>
            <div class="chart-wrapper"><canvas id="openAllChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR(ongoing) 趨勢 - 未解決問題 <span class="badge badge-internal">內部 JIRA</span></h2>
            <div class="chart-wrapper"><canvas id="openInternalChart"></canvas></div>
        </div>

        <div class="chart-container">
            <h2>�� MTTR(ongoing) 趨勢 - 未解決問題 <span class="badge badge-vendor">Vendor JIRA</span></h2>
            <div class="chart-wrapper"><canvas id="openVendorChart"></canvas></div>
        </div>

        <footer>
            <p>© 2025 Copyright by Vince. All rights reserved.</p>
            <p>MTTR 指標分析報告</p>
        </footer>
    </div>

    <script>
        const jiraSites = {{
            internal: '{data["jira_sites"]["internal"]}',
            vendor: '{data["jira_sites"]["vendor"]}'
        }};

        const mttrFilterIds = {{
            resolved: {{
                internal: '{MTTR_FILTERS["resolved"]["internal"] or ""}',
                vendor: '{MTTR_FILTERS["resolved"]["vendor"] or ""}'
            }},
            open: {{
                internal: '{MTTR_FILTERS["open"]["internal"] or ""}',
                vendor: '{MTTR_FILTERS["open"]["vendor"] or ""}'
            }}
        }};

        const weekDates = {{
            resolved_internal: {resolved_internal_dates},
            resolved_vendor: {resolved_vendor_dates},
            open_internal: {open_internal_dates},
            open_vendor: {open_vendor_dates}
        }};

        function openMttrWeekInJira(week, source, type) {{
            const site = jiraSites[source];
            const filterId = mttrFilterIds[type][source];
            if (!filterId) return;

            const dateKey = type + '_' + source;
            const weekData = weekDates[dateKey][week];
            if (!weekData) return;

            let jql = `filter=${{filterId}} AND created >= "${{weekData.start_date}} 00:00" AND created <= "${{weekData.end_date}} 23:59"`;
            const url = `https://${{site}}/issues/?jql=${{encodeURIComponent(jql)}}`;
            window.open(url, '_blank');
        }}

        function drawMttrChart(canvasId, labels, mttrData, overdueData, countData, source, type) {{
            const ctx = document.getElementById(canvasId).getContext('2d');
            const clickable = source !== null && type !== null;

            new Chart(ctx, {{
                type: 'line',
                data: {{
                    labels: labels,
                    datasets: [
                        {{
                            label: type === 'resolved' ? '平均 MTTR 天數' : '平均 MTTR(ongoing) 天數',
                            data: mttrData,
                            borderColor: '#667eea',
                            backgroundColor: 'rgba(102, 126, 234, 0.1)',
                            borderWidth: 3,
                            fill: true,
                            tension: 0.4,
                            yAxisID: 'y'
                        }},
                        {{
                            label: '平均 Overdue 天數',
                            data: overdueData,
                            borderColor: '#ff6b6b',
                            borderWidth: 3,
                            fill: false,
                            tension: 0.4,
                            yAxisID: 'y'
                        }},
                        {{
                            label: 'Issue 數量',
                            data: countData,
                            borderColor: '#51cf66',
                            borderWidth: 2,
                            borderDash: [5, 5],
                            fill: false,
                            tension: 0.4,
                            yAxisID: 'y1'
                        }}
                    ]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    interaction: {{ mode: 'index', intersect: false }},
                    onClick: clickable ? (event, elements) => {{
                        if (elements.length > 0) {{
                            const index = elements[0].index;
                            const week = labels[index];
                            openMttrWeekInJira(week, source, type);
                        }}
                    }} : undefined,
                    plugins: {{
                        legend: {{ display: true, position: 'top' }},
                        tooltip: {{
                            callbacks: {{
                                label: function(context) {{
                                    let label = context.dataset.label || '';
                                    if (label) label += ': ';
                                    if (context.parsed.y !== null) {{
                                        label += context.datasetIndex === 2 ? context.parsed.y + ' issues' : context.parsed.y.toFixed(2) + ' 天';
                                    }}
                                    return label;
                                }},
                                footer: clickable ? () => ['點擊查看 JIRA Issues'] : undefined
                            }}
                        }}
                    }},
                    scales: {{
                        y: {{
                            type: 'linear',
                            display: true,
                            position: 'left',
                            beginAtZero: true,
                            title: {{ display: true, text: '天數', color: '#667eea' }},
                            ticks: {{ color: '#667eea' }}
                        }},
                        y1: {{
                            type: 'linear',
                            display: true,
                            position: 'right',
                            beginAtZero: true,
                            title: {{ display: true, text: 'Issue 數量', color: '#51cf66' }},
                            ticks: {{ color: '#51cf66' }},
                            grid: {{ drawOnChartArea: false }}
                        }}
                    }}
                }}
            }});
        }}

        // 繪製所有圖表
        drawMttrChart('resolvedAllChart', {resolved_all_labels}, {resolved_all_mttr}, {resolved_all_overdue}, {resolved_all_count}, null, null);
        drawMttrChart('resolvedInternalChart', {resolved_internal_labels}, {resolved_internal_mttr}, {resolved_internal_overdue}, {resolved_internal_count}, 'internal', 'resolved');
        drawMttrChart('resolvedVendorChart', {resolved_vendor_labels}, {resolved_vendor_mttr}, {resolved_vendor_overdue}, {resolved_vendor_count}, 'vendor', 'resolved');
        drawMttrChart('openAllChart', {open_all_labels}, {open_all_mttr}, {open_all_overdue}, {open_all_count}, null, null);
        drawMttrChart('openInternalChart', {open_internal_labels}, {open_internal_mttr}, {open_internal_overdue}, {open_internal_count}, 'internal', 'open');
        drawMttrChart('openVendorChart', {open_vendor_labels}, {open_vendor_mttr}, {open_vendor_overdue}, {open_vendor_count}, 'vendor', 'open');

        console.log('✅ MTTR 圖表已載入');
    </script>
</body>
</html>
"""

        output = io.BytesIO(html_content.encode('utf-8'))
        output.seek(0)

        filename = f"mttr_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"

        return send_file(
            output,
            mimetype='text/html',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"❌ MTTR HTML 匯出失敗: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def get_local_ip():
    """取得本機 IP 位址"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        print(f"✅ 連接外部 IP: {ip}")
        return ip
    except Exception as e:
        print(f"❌ 失敗: {e}")
        return "127.0.0.1"

if __name__ == '__main__':
    # 從 .env 讀取設定，沒有才使用預設值
    host = os.getenv('FLASK_HOST', '0.0.0.0')
    port = int(os.getenv('FLASK_PORT', 5000))
    debug = os.getenv('FLASK_DEBUG', 'True').lower() == 'true'
    local_ip = get_local_ip()
    
    print("=" * 70)
    print("�� JIRA Degrade 分析系統 - 啟動中...")
    print("=" * 70)
    print()
    print("�� 系統資訊:")
    print(f"   • 版本: v2.1 (2025-11-18)")
    print(f"   • 作者: Vince")
    print()
    print("⚙️  設定資訊:")
    print(f"   • Flask Host: {host}")
    print(f"   • Flask Port: {port}")
    print(f"   • Debug Mode: {debug}")
    print(f"   • Cache TTL: {cache.ttl}秒")
    print()
    print("�� Degrade Filter IDs:")
    print(f"   • 內部 Degrade: {FILTERS['degrade']['internal']}")
    print(f"   • Vendor Degrade: {FILTERS['degrade']['vendor']}")
    print(f"   • 內部 Resolved: {FILTERS['resolved']['internal']}")
    print(f"   • Vendor Resolved: {FILTERS['resolved']['vendor']}")
    print()
    if MTTR_ENABLED:
        print("�� MTTR Filter IDs (已啟用):")
        print(f"   • 內部 Resolved: {MTTR_FILTERS['resolved']['internal']}")
        print(f"   • Vendor Resolved: {MTTR_FILTERS['resolved']['vendor']}")
        print(f"   • 內部 Open: {MTTR_FILTERS['open']['internal']}")
        print(f"   • Vendor Open: {MTTR_FILTERS['open']['vendor']}")
    else:
        print("�� MTTR 指標: 未啟用 (未設定 MTTR Filter IDs)")
    print()
    print("�� 功能說明:")
    print("   ✅ 統一從 .env 讀取所有設定")
    print("   ✅ Degrade issues 使用 created 日期")
    print("   ✅ Resolved issues 使用 created 日期")
    print("   ✅ 趨勢圖加入 CCC issue 數量線（雙 Y 軸）")
    print("   ✅ 週次日期範圍計算精確化")
    print("   ✅ 匯出 HTML 紅框連結可點擊")
    if MTTR_ENABLED:
        print("   ✅ MTTR 指標頁籤（已解決/未解決問題分析）")
    print()
    print("�� 伺服器位址:")
    print(f"   • 本機訪問: http://127.0.0.1:{port}")
    print(f"   • 區域網路訪問: http://{local_ip}:{port}")
    print()
    print("�� 提示:")
    print("   • 首次載入需要 30-60 秒")
    print("   • 按 Ctrl+C 停止服務")
    print("   • 查看 README.md 了解更多功能")
    print()
    print("=" * 70)
    print()
    
    app.run(debug=debug, host=host, port=port)